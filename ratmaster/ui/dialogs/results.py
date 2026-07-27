"""
ui/dialogs/results.py
=====================
Diálogos de resultados del cálculo BNCT.

ComponentDosesDialog
    Tabla con las componentes de dosis física por órgano:
    Boro, Neutrones Rápidos (Fstn), Térmicos (Thn) y Gamma.
    Para cada componente muestra Dmean, Dmax y la fracción (%)
    respecto al total físico del órgano.

ResultsDialog
    Ventana principal de resultados con:
      - Tabla de métricas dosimétricas (Dmax, Dmean, D95, D5, Dmin).
      - DVH combinado (todos los órganos).
      - Información del constraint limitante.
      - Botón de exportación a PDF (ReportLab).
      - Botón de acceso a ComponentDosesDialog.
"""

import io
from pathlib import Path
from datetime import datetime
import tempfile
import numpy as np
from matplotlib.figure import Figure          # para renderizar DVH offline en el PDF
from matplotlib.backends.backend_agg import FigureCanvasAgg  # backend sin pantalla
from matplotlib.backends.backend_qtagg import NavigationToolbar2QT  # zoom/pan en el DVH interactivo
from PySide6 import QtCore, QtWidgets, QtGui

from ratmaster.constants import ORG_ORDER, ORGAN_COLORS, _resolve_boro_protocol_name
from ratmaster.physics.dose_utils import build_dvh, dvh_extend_to_zero, achieved_value_for_constraint
from ratmaster.ui.formatters import format_value_uncertainty, format_scientific_value_uncertainty, format_value_uncertainty_separate, format_time, dose_axis_unit, format_boro_origin
from ratmaster.ui.canvas import MplCanvas

try:
    from reportlab.lib.pagesizes import A4
    from reportlab.platypus import (
        SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image as RLImage,
    )
    from reportlab.lib import colors
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.lib.units import cm
    HAVE_REPORTLAB = True
except Exception:
    HAVE_REPORTLAB = False

try:
    from ratmaster.physics.radiobio.models import compute_radiobio_report, RadiobioOrganParams
    RADIOBIO_OK = True
except ImportError:
    RADIOBIO_OK = False


# ─────────────────────────────────────────────────────────────────────────────
# Helper: copiar tabla como TSV para pegar en Excel
# ─────────────────────────────────────────────────────────────────────────────

def _install_copy_shortcut(table: QtWidgets.QTableWidget) -> None:
    """
    Instala Ctrl+C en una QTableWidget para copiar las celdas seleccionadas
    como texto tabulado (TSV), compatible con pegar directamente en Excel.
    Cada fila termina con \\n y las columnas se separan con \\t.
    """
    table.setSelectionMode(QtWidgets.QAbstractItemView.ExtendedSelection)

    def _copy():
        indexes = table.selectedIndexes()
        if not indexes:
            return
        rows = sorted({idx.row() for idx in indexes})
        cols = sorted({idx.column() for idx in indexes})
        lines = []
        for r in rows:
            row_data = []
            for c in cols:
                item = table.item(r, c)
                row_data.append(item.text() if item else "")
            lines.append("\t".join(row_data))
        QtWidgets.QApplication.clipboard().setText("\n".join(lines))

    shortcut = QtGui.QShortcut(QtGui.QKeySequence.Copy, table)
    shortcut.setContext(QtCore.Qt.WidgetShortcut)
    shortcut.activated.connect(_copy)


def _selectable_label(text: str) -> QtWidgets.QLabel:
    """Crea un QLabel cuyo texto puede seleccionarse y copiarse con el mouse."""
    lbl = QtWidgets.QLabel(text)
    lbl.setTextInteractionFlags(
        QtCore.Qt.TextSelectableByMouse | QtCore.Qt.TextSelectableByKeyboard
    )
    lbl.setCursor(QtGui.QCursor(QtCore.Qt.IBeamCursor))
    return lbl



# ═══════════════════════════════════════════════════════════════════════════
# MeshSegViewerDialog — visor interactivo de alineación SEG ↔ Mesh MCNP
# ═══════════════════════════════════════════════════════════════════════════

class MeshSegViewerDialog(QtWidgets.QDialog):
    """
    Visor 2D interactivo para verificar la alineación entre la segmentación
    (SEG) y las meshes del archivo meshtal de MCNP.

    Muestra cortes Axial / Coronal / Sagital con:
      - Fondo:   valores del tally seleccionado como heatmap (jet, escala lineal
                 o logarítmica).
      - Overlay: órganos del SEG con colores distintos (tab20) y transparencia
                 ajustable.
      - Borde:   rectángulo blanco punteado que marca el límite de la mesh en
                 el plano del corte — la anatomía útil debe quedar dentro.

    Uso:
        dlg = MeshSegViewerDialog(parent, segM, meshes,
                                  voxel_size_mm, origin_mesh_cm, lut)
        dlg.exec()

    Args:
        segM:           ndarray (nX, nY, nZ) int con etiquetas del SEG (ya
                        transformado con swap_YZ + rotate si corresponde).
        meshes:         dict de read_meshtal_all() — {tally_num: {...}}.
        voxel_size_mm:  (sx, sy, sz) tamaño de vóxel del SEG en mm.
        origin_mesh_cm: (ox, oy, oz) origen del SEG en coordenadas MCNP (cm).
        lut:            {label_int: nombre_organ} para la leyenda.
    """

    _ORIENT = {
        # (fix_axis, col_axis, row_axis, xlabel, ylabel, label)
        0: (2, 0, 1, "X [vox]", "Y [vox]", "Coronal (Z fijo)"),
        1: (1, 0, 2, "X [vox]", "Z [vox]", "Sagital (Y fijo)"),
        2: (0, 1, 2, "Y [vox]", "Z [vox]", "Axial   (X fijo)"),
    }

    _TALLY_LABELS = {
        14:  "Boro (14)",
        24:  "Térmicos (24)",
        34:  "Rápidos (34)",
        44:  "Gamma (44)",
        74:  "Flujo (74)",
    }

    def __init__(self, parent,
                 segM: np.ndarray,
                 meshes: dict,
                 voxel_size_mm,
                 origin_mesh_cm,
                 lut: dict):
        super().__init__(parent)
        self.setWindowTitle("Verificación de Alineación SEG ↔ Mesh — RatMaster")
        self.resize(980, 600)
        # self.setMinimumSize(720, 580)

        self.segM           = segM
        self.meshes         = meshes
        self.voxel_size_mm  = tuple(float(x) for x in voxel_size_mm)
        self.origin_mesh_cm = tuple(float(x) for x in origin_mesh_cm)
        self.lut            = lut

        # Estado de navegación
        self.orientation   = 0
        self.current_slice = segM.shape[2] // 2   # corte central axial
        self.seg_alpha     = 0.45
        self.log_scale     = True
        self.current_tally = next(iter(meshes), None)

        # Interpoladores (construidos una vez, usados en cada render)
        self._interps: dict = {}
        self._build_interps()

        # Colores por órgano (tab20, reproducibles)
        self._organ_colors: dict[int, tuple] = {}
        self._build_organ_colors()

        self._setup_ui()

        # Render inicial
        self._on_orientation_changed(0)

    # ── Pre-cómputo ───────────────────────────────────────────────────────

    def _build_interps(self):
        """Construye un RegularGridInterpolator por tally."""
        from scipy.interpolate import RegularGridInterpolator

        for tally, mesh in self.meshes.items():
            cx = mesh["centers"]["x"].copy()
            cy = mesh["centers"]["y"].copy()
            cz = mesh["centers"]["z"].copy()
            M  = mesh["Matrix"].astype(np.float64)

            # Garantizar ejes crecientes
            if cx.size >= 2 and cx[0] > cx[-1]:
                cx = cx[::-1]; M = np.flip(M, 0)
            if cy.size >= 2 and cy[0] > cy[-1]:
                cy = cy[::-1]; M = np.flip(M, 1)
            if cz.size >= 2 and cz[0] > cz[-1]:
                cz = cz[::-1]; M = np.flip(M, 2)

            self._interps[tally] = RegularGridInterpolator(
                (cx, cy, cz), M,
                method="linear",
                bounds_error=False,
                fill_value=0.0,
            )

    def _build_organ_colors(self):
        """Asigna color RGB único a cada label de órgano (tab20)."""
        import matplotlib.pyplot as _plt
        labels = sorted(int(l) for l in np.unique(self.segM) if l > 0)
        cmap   = _plt.get_cmap("tab20", max(len(labels), 1))
        for i, lab in enumerate(labels):
            r, g, b, _ = cmap(i % 20)
            self._organ_colors[lab] = (r, g, b)

    # ── UI ───────────────────────────────────────────────────────────────

    def _setup_ui(self):
        from matplotlib.backends.backend_qtagg import (
            FigureCanvasQTAgg, NavigationToolbar2QT,
        )
        from matplotlib.figure import Figure as MplFigure

        root = QtWidgets.QVBoxLayout(self)
        root.setContentsMargins(8, 6, 8, 6)
        root.setSpacing(4)

        # ── Barra superior ─────────────────────────────────────────────
        top = QtWidgets.QHBoxLayout()

        top.addWidget(QtWidgets.QLabel("<b>Tally:</b>"))
        self.tally_combo = QtWidgets.QComboBox()
        for t in self.meshes:
            label = self._TALLY_LABELS.get(t, f"Tally {t}")
            self.tally_combo.addItem(label, userData=t)
        top.addWidget(self.tally_combo)

        top.addSpacing(16)
        self.chk_log = QtWidgets.QCheckBox("Escala log")
        self.chk_log.setChecked(self.log_scale)
        top.addWidget(self.chk_log)

        top.addSpacing(16)
        top.addWidget(QtWidgets.QLabel("Opacidad SEG:"))
        self.sld_alpha = QtWidgets.QSlider(QtCore.Qt.Horizontal)
        self.sld_alpha.setRange(0, 100)
        self.sld_alpha.setValue(int(self.seg_alpha * 100))
        self.sld_alpha.setFixedWidth(100)
        self.lbl_alpha = QtWidgets.QLabel(f"{int(self.seg_alpha*100)}%")
        self.lbl_alpha.setFixedWidth(34)
        top.addWidget(self.sld_alpha)
        top.addWidget(self.lbl_alpha)

        top.addStretch()
        root.addLayout(top)

        # ── Canvas ─────────────────────────────────────────────────────
        self.fig    = MplFigure(figsize=(9.5, 6.5))
        self.canvas = FigureCanvasQTAgg(self.fig)
        self.canvas.setSizePolicy(
            QtWidgets.QSizePolicy.Expanding,
            QtWidgets.QSizePolicy.Expanding,
        )
        self.toolbar = NavigationToolbar2QT(self.canvas, self)
        root.addWidget(self.toolbar)
        root.addWidget(self.canvas)

        # ── Orientación ────────────────────────────────────────────────
        ort_row = QtWidgets.QHBoxLayout()
        ort_row.addWidget(QtWidgets.QLabel("<b>Vista:</b>"))
        self.orient_group = QtWidgets.QButtonGroup(self)
        for bid, lbl in enumerate(["Coronal (Z)", "Sagital (Y)", "Axial   (X)"]):
            rb = QtWidgets.QRadioButton(lbl)
            rb.setChecked(bid == 0)
            self.orient_group.addButton(rb, bid)
            ort_row.addWidget(rb)
        ort_row.addStretch()
        root.addLayout(ort_row)

        # ── Slider de corte ────────────────────────────────────────────
        sl_row = QtWidgets.QHBoxLayout()
        sl_row.addWidget(QtWidgets.QLabel("Corte:"))
        self.slice_slider = QtWidgets.QSlider(QtCore.Qt.Horizontal)
        self.slice_slider.setMinimum(0)
        self.slice_slider.setMaximum(0)
        self.slice_label  = QtWidgets.QLabel("0 / 0")
        self.slice_label.setMinimumWidth(64)
        sl_row.addWidget(self.slice_slider, stretch=1)
        sl_row.addWidget(self.slice_label)
        root.addLayout(sl_row)

        # ── Botón cerrar ───────────────────────────────────────────────
        hbtn = QtWidgets.QHBoxLayout()
        hbtn.addStretch()
        hbtn.addWidget(QtWidgets.QPushButton("Cerrar", clicked=self.accept))
        root.addLayout(hbtn)

        # ── Señales ────────────────────────────────────────────────────
        self.tally_combo.currentIndexChanged.connect(self._on_tally_changed)
        self.chk_log.toggled.connect(self._on_log_changed)
        self.sld_alpha.valueChanged.connect(self._on_alpha_changed)
        self.orient_group.idClicked.connect(self._on_orientation_changed)
        self.slice_slider.valueChanged.connect(self._on_slice_changed)

    # ── Renderizado ───────────────────────────────────────────────────

    def _render_and_display(self):
        """Redibuja la figura completa para el estado actual."""
        import matplotlib.pyplot as _plt
        from matplotlib.patches import Rectangle

        nX, nY, nZ = self.segM.shape
        vx, vy, vz  = (s / 10.0 for s in self.voxel_size_mm)   # mm → cm
        ox, oy, oz  = self.origin_mesh_cm
        s           = self.current_slice
        fix_ax, col_ax, row_ax, xlabel, ylabel, orient_name = \
            self._ORIENT[self.orientation]

        # ── Coordenadas físicas del corte ──────────────────────────────
        if self.orientation == 0:          # Axial: fix Z=s, show X(col) × Y(row)
            xi = (np.arange(nX) + 0.5) * vx + ox
            yi = (np.arange(nY) + 0.5) * vy + oy
            XX, YY = np.meshgrid(xi, yi, indexing="ij")   # (nX, nY)
            ZZ      = np.full_like(XX, (s + 0.5) * vz + oz)
            seg_2d  = self.segM[:, :, s]                   # (nX, nY)
            nrow, ncol = nY, nX

        elif self.orientation == 1:        # Coronal: fix Y=s, show X(col) × Z(row)
            xi = (np.arange(nX) + 0.5) * vx + ox
            zi = (np.arange(nZ) + 0.5) * vz + oz
            XX, ZZ = np.meshgrid(xi, zi, indexing="ij")   # (nX, nZ)
            YY      = np.full_like(XX, (s + 0.5) * vy + oy)
            seg_2d  = self.segM[:, s, :]                   # (nX, nZ)
            nrow, ncol = nZ, nX

        else:                              # Sagital: fix X=s, show Y(col) × Z(row)
            yi = (np.arange(nY) + 0.5) * vy + oy
            zi = (np.arange(nZ) + 0.5) * vz + oz
            YY, ZZ = np.meshgrid(yi, zi, indexing="ij")   # (nY, nZ)
            XX      = np.full_like(YY, (s + 0.5) * vx + ox)
            seg_2d  = self.segM[s, :, :]                   # (nY, nZ)
            nrow, ncol = nZ, nY

        # Apilar coords para interpolación: (N, 3) en orden (X, Y, Z)
        pts = np.column_stack([XX.ravel(), YY.ravel(), ZZ.ravel()])

        # ── Valores de la mesh interpolados en la grilla del SEG ───────
        interp = self._interps.get(self.current_tally)
        if interp is not None:
            vals = interp(pts).reshape(XX.shape)
        else:
            vals = np.zeros(XX.shape)

        # Transponer a (nrow, ncol) para imshow (col_ax → cols, row_ax → rows)
        mesh_img = vals.T                        # (nY,nX) / (nZ,nX) / (nZ,nY)
        seg_disp = seg_2d.T                      # idem

        # Escala logarítmica opcional
        with np.errstate(divide="ignore", invalid="ignore"):
            if self.log_scale and np.any(mesh_img > 0):
                min_pos = float(np.nanmin(mesh_img[mesh_img > 0]))
                floor   = min_pos * 1e-2
                mesh_plot = np.where(mesh_img > 0,
                                     np.log10(np.maximum(mesh_img, floor)),
                                     np.log10(floor))
                cbar_label = "log₁₀(tally)"
            else:
                mesh_plot  = mesh_img.copy()
                cbar_label = "tally"

        vmin = float(np.nanmin(mesh_plot)) if np.any(np.isfinite(mesh_plot)) else 0.0
        vmax = float(np.nanmax(mesh_plot)) if np.any(np.isfinite(mesh_plot)) else 1.0
        if vmin == vmax:
            vmax = vmin + 1.0

        # ── RGBA de órganos ────────────────────────────────────────────
        seg_rgba = np.zeros((nrow, ncol, 4), dtype=np.float32)
        for lab, (r, g, b) in self._organ_colors.items():
            mask = seg_disp == lab
            seg_rgba[mask, 0] = r
            seg_rgba[mask, 1] = g
            seg_rgba[mask, 2] = b
            seg_rgba[mask, 3] = self.seg_alpha

        # ── Figura ─────────────────────────────────────────────────────
        self.fig.clear()
        ax      = self.fig.add_axes([0.05, 0.06, 0.80, 0.88])
        cbar_ax = self.fig.add_axes([0.87, 0.06, 0.025, 0.88])

        im = ax.imshow(mesh_plot, cmap="jet",
                       vmin=vmin, vmax=vmax,
                       origin="lower", aspect="auto", interpolation="nearest")
        ax.imshow(seg_rgba,
                  origin="lower", aspect="auto", interpolation="nearest")

        # ── Rectángulo de límite de la mesh ────────────────────────────
        mesh = self.meshes.get(self.current_tally, {})
        cx_m = mesh.get("centers", {}).get("x")
        cy_m = mesh.get("centers", {}).get("y")
        cz_m = mesh.get("centers", {}).get("z")
        if cx_m is not None:
            def _to_pix(phys, origin, step):
                """Convierte coordenada física (cm) a píxel del SEG."""
                return (np.array(phys) - origin) / step - 0.5

            if self.orientation == 0:     # Axial: X→col, Y→row
                c0, c1 = _to_pix([cx_m.min(), cx_m.max()], ox, vx)
                r0, r1 = _to_pix([cy_m.min(), cy_m.max()], oy, vy)
            elif self.orientation == 1:   # Coronal: X→col, Z→row
                c0, c1 = _to_pix([cx_m.min(), cx_m.max()], ox, vx)
                r0, r1 = _to_pix([cz_m.min(), cz_m.max()], oz, vz)
            else:                          # Sagital: Y→col, Z→row
                c0, c1 = _to_pix([cy_m.min(), cy_m.max()], oy, vy)
                r0, r1 = _to_pix([cz_m.min(), cz_m.max()], oz, vz)

            rect = Rectangle(
                (c0, r0), c1 - c0, r1 - r0,
                linewidth=1.8, edgecolor="white",
                linestyle="--", facecolor="none", alpha=0.9, zorder=5,
            )
            ax.add_patch(rect)
            ax.text(c0 + 1, r1 - 1, "límite mesh",
                    color="white", fontsize=7, va="top", zorder=6,
                    bbox=dict(boxstyle="round,pad=0.1", fc="black", alpha=0.4))

        # ── Leyenda de órganos ─────────────────────────────────────────
        from matplotlib.patches import Patch
        handles = [
            Patch(facecolor=color, edgecolor="none",
                  label=self.lut.get(lab, f"lab {lab}"))
            for lab, color in self._organ_colors.items()
            if np.any(seg_disp == lab)
        ]
        if handles:
            ax.legend(handles=handles, loc="lower right",
                      fontsize=6.5, framealpha=0.6,
                      ncol=max(1, len(handles) // 10))

        # ── Colorbar y etiquetas ───────────────────────────────────────
        cbar = self.fig.colorbar(im, cax=cbar_ax)
        cbar.set_label(cbar_label, fontsize=8)
        cbar.ax.tick_params(labelsize=7)

        n_slices = self.segM.shape[fix_ax]
        ax.set_xlabel(xlabel, fontsize=9)
        ax.set_ylabel(ylabel, fontsize=9)
        tally_lbl = self._TALLY_LABELS.get(self.current_tally,
                                            f"Tally {self.current_tally}")
        ax.set_title(
            f"{orient_name}  —  corte {s} / {n_slices - 1}"
            f"  —  {tally_lbl}",
            fontsize=10,
        )
        ax.tick_params(labelsize=8)
        self.canvas.draw()

    # ── Slots ─────────────────────────────────────────────────────────

    def _on_tally_changed(self, _=None):
        self.current_tally = self.tally_combo.currentData()
        self._render_and_display()

    def _on_log_changed(self, checked: bool):
        self.log_scale = checked
        self._render_and_display()

    def _on_alpha_changed(self, val: int):
        self.seg_alpha = val / 100.0
        self.lbl_alpha.setText(f"{val}%")
        self._render_and_display()

    def _on_orientation_changed(self, bid: int):
        self.orientation = bid
        n = self.segM.shape[self._ORIENT[bid][0]]
        mid = n // 2
        self.slice_slider.blockSignals(True)
        self.slice_slider.setMaximum(n - 1)
        self.slice_slider.setValue(mid)
        self.slice_slider.blockSignals(False)
        self.current_slice = mid
        self.slice_label.setText(f"{mid} / {n - 1}")
        self._render_and_display()

    def _on_slice_changed(self, val: int):
        self.current_slice = val
        n = self.slice_slider.maximum() + 1
        self.slice_label.setText(f"{val} / {n - 1}")
        self._render_and_display()


# ═══════════════════════════════════════════════════════════════════════════
# Helper: construcción de datos espaciales para el visor 2D
# ═══════════════════════════════════════════════════════════════════════════

def build_viz_data(organ_dose: dict, report: dict,
                   segM: np.ndarray, voxel_size_mm) -> dict:
    """
    Construye el dict `viz_data` necesario para DoseViewerDialog.

    Debe llamarse justo después de `calcular_dosis_por_organo_trilineal` y
    `compute_bnct`, antes de que se descarte `organ_dose` y `segM`.

    Args:
        organ_dose:    salida de calcular_dosis_por_organo_trilineal — contiene
                       "original_indices" (N_total,3) y "valid_mask" (N_total,)
                       por órgano.
        report:        salida de compute_bnct — contiene PhysVoxel, BioVoxel,
                       IsoVoxel por órgano.
        segM:          segmentación (X,Y,Z) int con etiquetas de órgano.
                       Se usa solo para extraer la máscara de cuerpo (label > 0).
        voxel_size_mm: (sx, sy, sz) tamaño de vóxel en mm.

    Returns:
        viz_data: dict con claves:
            "seg_shape"           — tuple (X,Y,Z)
            "voxel_size_mm"       — tuple (sx,sy,sz)
            "body_mask_indices"   — ndarray (N_body,3) ijk de vóxeles con label>0
            "organ_valid_indices" — {organ: ndarray (N_valid,3)} índices ijk
                                    de vóxeles válidos (con datos MCNP) por órgano
            "phys_dose"           — {organ: ndarray (N_valid,)} dosis física
            "bio_dose"            — {organ: ndarray (N_valid,)} dosis biológica
            "iso_dose"            — {organ: ndarray (N_valid,)} dosis isoefectiva

    Nota:
        body_mask_indices se usa como fondo del visor:
          - Fuera del cuerpo (label == 0) → negro
          - Interior del cuerpo (label > 0) → gris
        Esta representación evita guardar la segmentación completa.
    """
    seg_shape = segM.shape                              # (X, Y, Z)
    body_ijk  = np.argwhere(segM > 0).astype(np.int32) # (N_body, 3)

    organ_valid_indices: dict[str, np.ndarray] = {}
    for organ, data in organ_dose.items():
        orig_idx  = data.get("original_indices")
        valid_mask = data.get("valid_mask")
        if orig_idx is not None and valid_mask is not None:
            organ_valid_indices[organ] = orig_idx[valid_mask].astype(np.int32)

    def _extract(voxkey: str) -> dict[str, np.ndarray]:
        out = {}
        vmap = report.get(voxkey, {})
        for organ in organ_valid_indices:
            raw = vmap.get(organ)
            if raw is not None:
                out[organ] = np.asarray(raw, dtype=np.float64)
        return out

    return {
        "seg_shape":             tuple(int(x) for x in seg_shape),
        "voxel_size_mm":         tuple(float(x) for x in voxel_size_mm),
        "body_mask_indices":     body_ijk,
        "organ_valid_indices":   organ_valid_indices,
        "phys_dose":             _extract("PhysVoxel"),
        "bio_dose":              _extract("BioVoxel"),
        "iso_dose":              _extract("IsoVoxel"),
    }


# ═══════════════════════════════════════════════════════════════════════════
# DoseViewerDialog — visor 2D de mapas de dosis por órgano
# ═══════════════════════════════════════════════════════════════════════════

class DoseViewerDialog(QtWidgets.QDialog):
    """
    Visor 2D interactivo de dosis por vóxel superpuesta a la segmentación.

    Muestra cortes 2D en tres orientaciones (Axial / Coronal / Sagital)
    con un slider para navegar por los cortes.

    Fondo de pantalla:
        Negro  → exterior del cuerpo (sin etiqueta en la segmentación)
        Gris   → interior del cuerpo (cualquier órgano registrado)

    Superposición de dosis:
        Colormap 'jet' (azul=bajo, rojo=alto) con transparencia α=0.82.
        Solo sobre los vóxeles válidos (con datos MCNP) del órgano elegido.

    Construir `viz_data` con `build_viz_data()` y pasarlo como argumento.
    """

    _CMAP = "jet"

    _ORIENT = {
        #  (fix_axis, col_axis, row_axis, xlabel, ylabel, label)
        0: (2, 0, 1, "X [vox]", "Y [vox]", "Coronal (Z fijo)"),
        1: (1, 0, 2, "X [vox]", "Z [vox]", "Sagital (Y fijo)"),
        2: (0, 1, 2, "Y [vox]", "Z [vox]", "Axial   (X fijo)"),
    }

    def __init__(self, parent, viz_data: dict, report: dict,
                 use_bio: bool = False, is_isoe: bool = False):
        super().__init__(parent)
        self.setWindowTitle("Visualizador de Dosis — RatMaster")
        self.resize(960, 600)
        self.setMinimumSize(720, 580)

        self.viz_data  = viz_data
        self.report    = report
        self.use_bio   = use_bio
        self.is_isoe   = is_isoe

        # Estado de navegación
        self.orientation   = 0
        self.current_slice = 0
        self.current_organ = ""
        self.vmin          = 0.0
        self.vmax          = 1.0

        # Caché de arrays de dosis (se invalida al cambiar órgano o tipo)
        self._dose_arr_cache: dict[str, np.ndarray] = {}

        # Sets de dosis disponibles (nombre -> {organ: raw list})
        self._dose_sets: dict[str, dict] = {}
        for key, label in (("phys_dose", "Física"),
                            ("bio_dose",  "Biológica"),
                            ("iso_dose",  "Isoefectiva")):
            ds = viz_data.get(key, {})
            if ds:
                self._dose_sets[label] = ds

        if not self._dose_sets:
            # Fallback: intentar desde el report directamente
            for key, label in (("PhysVoxel", "Física"),
                                ("BioVoxel",  "Biológica"),
                                ("IsoVoxel",  "Isoefectiva")):
                ds = report.get(key, {})
                if ds:
                    self._dose_sets[label] = ds

        # Tipo de dosis seleccionado por defecto
        if self.is_isoe and "Isoefectiva" in self._dose_sets:
            self._dose_key = "Isoefectiva"
        elif self.use_bio and "Biológica" in self._dose_sets:
            self._dose_key = "Biológica"
        elif "Física" in self._dose_sets:
            self._dose_key = "Física"
        else:
            self._dose_key = next(iter(self._dose_sets), "Física")

        # ── Construir mapeo de órganos ────────────────────────────────────────
        # Las claves en organ_valid_indices vienen del LUT del SEG (pueden diferir
        # en capitalización o tener sufijos). Las claves en phys_dose vienen de
        # ORG_ORDER/compute_bnct. Se usan con normalización lowercase+stripped.
        #
        # _organ_key_map[display_name] = {"spatial": key_en_viz_data,
        #                                  "dose":    key_en_phys/bio/iso_dose}
        # Solo se incluyen órganos presentes en AMBOS (intersección).
        self._organ_key_map = self._build_organ_key_map(viz_data)

        self._setup_ui()

    # ── Matching de claves espaciales ↔ dosis ────────────────────────────

    @staticmethod
    def _norm_lower(s: str) -> str:
        """Normaliza para matching: minúsculas, sin acentos, solo alfanumérico."""
        import unicodedata as _ud, re as _re
        s = _ud.normalize("NFKD", str(s))
        s = "".join(c for c in s if not _ud.combining(c))
        return _re.sub(r"[^a-z0-9]", "", s.lower())

    def _build_organ_key_map(self, viz_data: dict) -> dict:
        """
        Construye el mapeo display_name → {spatial, dose} haciendo match
        normalizado entre organ_valid_indices y las claves de dosis disponibles.

        Reglas:
          1. Solo se incluyen órganos presentes en AMBOS lados (intersección).
          2. El matching es case-insensitive y sin caracteres especiales.
          3. Órganos no canónicos del SEG (sin dosis) quedan fuera del combo.
          4. Cualquier label que contenga "pulmon" en su nombre normalizado
             se combina en "PulmonTotal" (no depende de nombres exactos en el SEG).
          5. PulmonIzq / PulmonDer individuales se saltan del lado dosis.
          6. Tumor sin label propio hereda los índices de PulmonTotal.
        """
        nl = self._norm_lower

        # Trabajar sobre copia mutable para poder agregar PulmonTotal sintético
        spatial_raw: dict = dict(viz_data.get("organ_valid_indices", {}))

        # ── Crear PulmonTotal combinando todos los labels que contengan "pulmon" ──
        # Aplica independientemente de cómo se llamen en el SEG:
        # "PulmonIzq", "Pulmon_Der", "LungLeft", "PulmonD", etc.
        if not any(nl(k) == "pulmontotal" for k in spatial_raw):
            lung_parts = [v for k, v in spatial_raw.items() if "pulmon" in nl(k)]
            if lung_parts:
                combined = np.concatenate(lung_parts, axis=0)
                spatial_raw["PulmonTotal"] = combined
                # Persistir para que _build_slice_images lo encuentre después
                viz_data["organ_valid_indices"]["PulmonTotal"] = combined

        # Mapa norm → clave real en spatial (primer match gana)
        spatial_norm: dict[str, str] = {}
        for k in spatial_raw:
            spatial_norm.setdefault(nl(k), k)

        # Mapa norm → clave real en cualquier dose set (primer match gana)
        dose_norm: dict[str, str] = {}
        for ds in self._dose_sets.values():
            for k in ds:
                dose_norm.setdefault(nl(k), k)

        # Norms de pulmones individuales que NO deben aparecer en el combo
        _LUNG_IND = {"pulmonizq", "pulmonder", "pulmonizquierdo",
                     "pulmonleft", "pulmonright"}

        key_map: dict[str, dict] = {}

        for norm, dose_key in dose_norm.items():
            # Saltar pulmones individuales del lado dosis
            if norm in _LUNG_IND:
                continue

            if norm in spatial_norm:
                # Match directo normalizado (case-insensitive, sin especiales)
                key_map[dose_key] = {
                    "spatial": spatial_norm[norm],
                    "dose":    dose_key,
                }
            elif nl(dose_key) == "tumor":
                # Tumor comparte ubicación espacial con el pulmón en modelos
                # de tumor pulmonar de rata (vectors["Tumor"] = vectors["PulmonTotal"])
                pt_norm = nl("PulmonTotal")
                if pt_norm in spatial_norm:
                    key_map[dose_key] = {
                        "spatial": spatial_norm[pt_norm],
                        "dose":    dose_key,
                    }

        return key_map

        # Inicializar con primer órgano disponible en orientación Axial
        if self.organ_combo.count() > 0:
            self.organ_combo.setCurrentIndex(0)
            self.current_organ = self.organ_combo.currentText()
            self._on_orientation_changed(0)   # actualiza slider y redibujar
        else:
            self._render_and_display()

    # ── Construcción de UI ────────────────────────────────────────────────

    def _setup_ui(self):
        from matplotlib.backends.backend_qtagg import (
            FigureCanvasQTAgg, NavigationToolbar2QT,
        )
        from matplotlib.figure import Figure as MplFigure

        root = QtWidgets.QVBoxLayout(self)
        root.setContentsMargins(8, 8, 8, 8)
        root.setSpacing(5)

        # ── Barra superior ────────────────────────────────────────────────
        top = QtWidgets.QHBoxLayout()

        top.addWidget(QtWidgets.QLabel("<b>Órgano:</b>"))
        self.organ_combo = QtWidgets.QComboBox()
        self.organ_combo.addItems(sorted(self._organ_key_map.keys()))
        self.organ_combo.setMinimumWidth(170)
        top.addWidget(self.organ_combo)

        top.addSpacing(14)
        top.addWidget(QtWidgets.QLabel("<b>Dosis:</b>"))
        self.dose_combo = QtWidgets.QComboBox()
        self.dose_combo.addItems(list(self._dose_sets.keys()))
        self.dose_combo.setCurrentText(self._dose_key)
        self.dose_combo.setMinimumWidth(120)
        top.addWidget(self.dose_combo)

        top.addSpacing(14)
        top.addWidget(QtWidgets.QLabel("Min:"))
        self.spin_vmin = QtWidgets.QDoubleSpinBox()
        self.spin_vmin.setRange(0.0, 1e9)
        self.spin_vmin.setDecimals(4)
        self.spin_vmin.setSingleStep(0.001)
        self.spin_vmin.setFixedWidth(88)
        top.addWidget(self.spin_vmin)

        top.addWidget(QtWidgets.QLabel("Max:"))
        self.spin_vmax = QtWidgets.QDoubleSpinBox()
        self.spin_vmax.setRange(0.0, 1e9)
        self.spin_vmax.setDecimals(4)
        self.spin_vmax.setSingleStep(0.001)
        self.spin_vmax.setFixedWidth(88)
        top.addWidget(self.spin_vmax)

        self.btn_auto = QtWidgets.QPushButton("Auto")
        self.btn_auto.setFixedWidth(48)
        self.btn_auto.setToolTip("Escala al máximo del órgano actual")
        top.addWidget(self.btn_auto)

        top.addStretch()
        root.addLayout(top)

        # ── Canvas matplotlib ─────────────────────────────────────────────
        self.fig    = MplFigure(figsize=(9, 6))
        self.canvas = FigureCanvasQTAgg(self.fig)
        self.canvas.setSizePolicy(
            QtWidgets.QSizePolicy.Expanding,
            QtWidgets.QSizePolicy.Expanding,
        )
        self.toolbar = NavigationToolbar2QT(self.canvas, self)
        root.addWidget(self.toolbar)
        root.addWidget(self.canvas)

        # ── Botones de orientación ────────────────────────────────────────
        orient_row = QtWidgets.QHBoxLayout()
        orient_row.addWidget(QtWidgets.QLabel("<b>Vista:</b>"))
        self.orient_group = QtWidgets.QButtonGroup(self)
        orient_row.addSpacing(4)
        for bid, label in enumerate(
            ["Coronal (Z)", "Sagital (Y)", "Axial   (X)"]
        ):
            rb = QtWidgets.QRadioButton(label)
            rb.setChecked(bid == 0)
            self.orient_group.addButton(rb, bid)
            orient_row.addWidget(rb)
        orient_row.addStretch()
        root.addLayout(orient_row)

        # ── Slider de corte ───────────────────────────────────────────────
        slice_row = QtWidgets.QHBoxLayout()
        slice_row.addWidget(QtWidgets.QLabel("Corte:"))
        self.slice_slider = QtWidgets.QSlider(QtCore.Qt.Horizontal)
        self.slice_slider.setMinimum(0)
        self.slice_slider.setMaximum(0)
        self.slice_slider.setValue(0)
        self.slice_slider.setTickPosition(QtWidgets.QSlider.TicksBelow)
        self.slice_slider.setTickInterval(10)
        self.slice_label = QtWidgets.QLabel("0 / 0")
        self.slice_label.setMinimumWidth(72)
        self.slice_label.setAlignment(QtCore.Qt.AlignLeft | QtCore.Qt.AlignVCenter)
        self.n_vox_label = QtWidgets.QLabel("")
        self.n_vox_label.setStyleSheet("color: #546E7A; font-size: 9pt;")
        slice_row.addWidget(self.slice_slider, stretch=1)
        slice_row.addWidget(self.slice_label)
        slice_row.addSpacing(16)
        slice_row.addWidget(self.n_vox_label)
        root.addLayout(slice_row)

        # ── Botón cerrar ──────────────────────────────────────────────────
        hbtn = QtWidgets.QHBoxLayout()
        hbtn.addStretch()
        btn_close = QtWidgets.QPushButton("Cerrar")
        btn_close.clicked.connect(self.accept)
        hbtn.addWidget(btn_close)
        root.addLayout(hbtn)

        # ── Señales ───────────────────────────────────────────────────────
        self.organ_combo.currentIndexChanged.connect(self._on_organ_changed)
        self.dose_combo.currentTextChanged.connect(self._on_dose_type_changed)
        self.orient_group.idClicked.connect(self._on_orientation_changed)
        self.slice_slider.valueChanged.connect(self._on_slice_changed)
        self.spin_vmin.valueChanged.connect(self._on_clim_changed)
        self.spin_vmax.valueChanged.connect(self._on_clim_changed)
        self.btn_auto.clicked.connect(self._autoscale)
        # ── Render inicial ────────────────────────────────────────────────
        if self.organ_combo.count() > 0:
            self.current_organ = self.organ_combo.itemText(0)
        # Inicializar el slider de cortes con la orientación por defecto
        n = max(self._slice_count(0), 1)
        mid = n // 2
        self.slice_slider.blockSignals(True)
        self.slice_slider.setMaximum(n - 1)
        self.slice_slider.setValue(mid)
        self.slice_slider.blockSignals(False)
        self.current_slice = mid
        self.slice_label.setText(f"{mid} / {n - 1}")
        # Autoscale ajusta vmin/vmax y llama a _render_and_display()
        self._autoscale()

    # ── Lógica de datos ──────────────────────────────────────────────────

    def _get_dose_arr(self, organ: str) -> np.ndarray | None:
        """Array de dosis del órgano en el set activo, con caché."""
        cache_key = f"{self._dose_key}::{organ}"
        if cache_key not in self._dose_arr_cache:
            raw = self._dose_sets.get(self._dose_key, {}).get(organ)
            if raw is None:
                return None
            self._dose_arr_cache[cache_key] = np.asarray(raw, dtype=np.float64)
        return self._dose_arr_cache[cache_key]

    def _slice_count(self, orientation: int) -> int:
        """Número de cortes en la orientación dada."""
        seg_shape = self.viz_data.get("seg_shape", (1, 1, 1))
        fix_ax = self._ORIENT[orientation][0]
        return seg_shape[fix_ax]

    def _autoscale(self):
        """Ajusta vmin/vmax al rango real del órgano actual."""
        dv = self._get_dose_arr(self.current_organ)
        if dv is None or dv.size == 0:
            vmin, vmax = 0.0, 1.0
        else:
            vmin = float(np.nanmin(dv))
            vmax = float(np.nanmax(dv))
            if vmin == vmax:
                vmax = vmin + 1.0

        self.vmin = vmin
        self.vmax = vmax
        for spin, val in ((self.spin_vmin, vmin), (self.spin_vmax, vmax)):
            spin.blockSignals(True)
            spin.setValue(val)
            spin.blockSignals(False)

        self._render_and_display()

    # ── Renderizado ──────────────────────────────────────────────────────

    def _build_slice_images(self, orientation: int, slice_idx: int):
        """
        Construye las dos imágenes para el corte solicitado:
            bg_2d:    float32 (nrow, ncol) — 0.0=negro, 0.35=gris-cuerpo
            dose_2d:  float64 (nrow, ncol) — dosis o NaN fuera del órgano
        """
        seg_shape  = self.viz_data.get("seg_shape", (1, 1, 1))
        body_ijk   = self.viz_data.get("body_mask_indices")
        org        = self.current_organ

        # Resolver la clave espacial correcta (puede diferir del display name)
        _km        = self._organ_key_map.get(org, {})
        spatial_key = _km.get("spatial", org)
        oi         = self.viz_data.get("organ_valid_indices", {}).get(spatial_key)
        dv         = self._get_dose_arr(org)

        fix_ax, col_ax, row_ax = self._ORIENT[orientation][:3]
        nrow = seg_shape[row_ax]
        ncol = seg_shape[col_ax]

        # ── Fondo (negro / gris-cuerpo) ───────────────────────────────────
        bg = np.zeros((nrow, ncol), dtype=np.float32)
        if body_ijk is not None and body_ijk.shape[0] > 0:
            sel = body_ijk[body_ijk[:, fix_ax] == slice_idx]
            if sel.shape[0] > 0:
                r = sel[:, row_ax].clip(0, nrow - 1)
                c = sel[:, col_ax].clip(0, ncol - 1)
                bg[r, c] = 0.35

        # ── Mapa de dosis (NaN donde no hay datos del órgano) ─────────────
        dose_map = np.full((nrow, ncol), np.nan, dtype=np.float64)
        n_vox_in_slice = 0
        if (oi is not None and dv is not None
                and oi.shape[0] > 0 and oi.shape[0] == dv.shape[0]):
            mask = oi[:, fix_ax] == slice_idx
            oi_s = oi[mask]
            dv_s = dv[mask]
            n_vox_in_slice = int(oi_s.shape[0])
            if n_vox_in_slice > 0:
                r = oi_s[:, row_ax].clip(0, nrow - 1)
                c = oi_s[:, col_ax].clip(0, ncol - 1)
                dose_map[r, c] = dv_s

        return bg, dose_map, n_vox_in_slice

    def _render_and_display(self):
        """Redibuja la figura completa con el estado actual."""
        bg, dose_map, n_vox = self._build_slice_images(
            self.orientation, self.current_slice
        )

        # Actualizar etiqueta de vóxeles en este corte
        if n_vox > 0:
            self.n_vox_label.setText(f"{n_vox} vóxeles del órgano en este corte")
        else:
            self.n_vox_label.setText("Sin vóxeles del órgano en este corte")

        # ── Reconstruir figura ────────────────────────────────────────────
        self.fig.clear()

        # Determinar si hay datos de dosis visibles en este corte
        masked_dose = np.ma.masked_invalid(dose_map)
        has_dose = bool(masked_dose.count() > 0)

        # Layout: axes principal + axes colorbar (solo si hay dosis)
        if has_dose:
            ax_main = self.fig.add_axes([0.07, 0.06, 0.78, 0.88])
            cbar_ax = self.fig.add_axes([0.87, 0.06, 0.025, 0.88])
        else:
            ax_main = self.fig.add_axes([0.07, 0.06, 0.88, 0.88])

        # Fondo gris/negro
        ax_main.imshow(
            bg,
            cmap="gray", vmin=0.0, vmax=1.0,
            origin="lower", aspect="auto", interpolation="nearest",
        )

        # Mapa de dosis superpuesto
        if has_dose:
            im = ax_main.imshow(
                masked_dose,
                cmap=self._CMAP,
                vmin=self.vmin,
                vmax=max(self.vmax, self.vmin + 1e-10),
                origin="lower", aspect="auto", interpolation="nearest",
                alpha=0.82,
            )
            cbar = self.fig.colorbar(im, cax=cbar_ax)
            unit_map = {
                "Física":      "Gy",
                "Biológica":   "Gy(RBE)",
                "Isoefectiva": "Gy(IsoE)",
            }
            cbar.set_label(
                unit_map.get(self._dose_key, "Gy"),
                fontsize=9,
            )
            cbar.ax.tick_params(labelsize=8)

        # Etiquetas
        fix_ax, col_ax, row_ax, xlabel, ylabel, orient_name = self._ORIENT[self.orientation]
        n_total = self._slice_count(self.orientation)
        ax_main.set_xlabel(xlabel, fontsize=9)
        ax_main.set_ylabel(ylabel, fontsize=9)
        ax_main.set_title(
            f"{orient_name}  —  corte {self.current_slice} / {n_total - 1}"
            f"  —  {self.current_organ or '(ningún órgano)'}",
            fontsize=10,
        )
        ax_main.tick_params(labelsize=8)

        self.canvas.draw()

    # ── Slots ────────────────────────────────────────────────────────────

    def _on_organ_changed(self, _idx: int = 0):
        self.current_organ = self.organ_combo.currentText()
        self._dose_arr_cache = {}
        # autoscale redibuja
        self._autoscale()

    def _on_dose_type_changed(self, text: str):
        self._dose_key = text
        self._dose_arr_cache = {}
        self._autoscale()

    def _on_orientation_changed(self, bid: int):
        self.orientation = bid
        n = self._slice_count(bid)
        n = max(n, 1)
        mid = n // 2

        self.slice_slider.blockSignals(True)
        self.slice_slider.setMaximum(n - 1)
        self.slice_slider.setValue(mid)
        self.slice_slider.blockSignals(False)

        self.current_slice = mid
        self.slice_label.setText(f"{mid} / {n - 1}")
        self._render_and_display()

    def _on_slice_changed(self, val: int):
        self.current_slice = val
        n = self.slice_slider.maximum() + 1
        self.slice_label.setText(f"{val} / {n - 1}")
        self._render_and_display()

    def _on_clim_changed(self, _=None):
        self.vmin = self.spin_vmin.value()
        self.vmax = self.spin_vmax.value()
        if self.vmax <= self.vmin:
            self.vmax = self.vmin + 1e-6
        self._render_and_display()


# ───────────────────────────────────────────────────────────────────────────
class CalculationParamsDialog(QtWidgets.QDialog):
    """
    Muestra todos los parámetros usados en el cálculo en un único scroll vertical.

    Secciones (siempre visibles a menos que se indique):
      1. Condiciones generales   — fecha, modo, protocolo, tiempo, flujo
      2. Boro / CBE / RBE        — tabla por órgano con B±σ, CBE, RBE
      3. Constraints usadas      — solo si modo constraints
      4. SPND / Corrientes       — solo si se calculó flujo desde corrientes
      5. Parámetros IsoE         — solo si modo IsoE (α/β, t0 por componente)
    """

    # Estilos de tabla reutilizables
    _HDR_COLOR  = "#455A64"
    _HDR_TEXT   = "#FFFFFF"
    _ROW_EVEN   = "#F5F7FA"
    _ROW_ODD    = "#FFFFFF"
    _BORDER     = "#CFD8DC"

    def __init__(self, parent, report: dict, dose_mode_text: str,
                 use_bio: bool = False, is_isoe: bool = False):
        super().__init__(parent)
        self.setWindowTitle("Parámetros del cálculo")
        self.resize(750, 600)

        meta   = report.get("meta", {}) or {}
        params = report.get("ParamsUsed", {}) or {}
        unit   = dose_axis_unit(use_bio, is_isoe)

        # ── Layout raíz con scroll ────────────────────────────────────────
        root = QtWidgets.QVBoxLayout(self)
        root.setContentsMargins(0, 0, 0, 8)

        scroll = QtWidgets.QScrollArea()
        scroll.setWidgetResizable(True)
        scroll.setFrameShape(QtWidgets.QFrame.NoFrame)
        root.addWidget(scroll)

        container = QtWidgets.QWidget()
        vlay = QtWidgets.QVBoxLayout(container)
        vlay.setSpacing(14)
        vlay.setContentsMargins(16, 16, 16, 8)
        scroll.setWidget(container)

        # ── Helper: crear GroupBox con QVBoxLayout ────────────────────────
        def _group(title: str) -> tuple:
            gb = QtWidgets.QGroupBox(title)
            gl = QtWidgets.QVBoxLayout(gb)
            gl.setSpacing(4)
            vlay.addWidget(gb)
            return gb, gl

        # ── Helper: tabla simple ──────────────────────────────────────────
        def _table(parent_layout, headers: list, rows: list,
                   col_stretch: list | None = None):
            tbl = QtWidgets.QTableWidget(len(rows), len(headers))
            tbl.setHorizontalHeaderLabels(headers)
            tbl.verticalHeader().setVisible(False)
            tbl.setShowGrid(True)
            tbl.setEditTriggers(QtWidgets.QAbstractItemView.NoEditTriggers)
            tbl.setSelectionMode(QtWidgets.QAbstractItemView.SingleSelection)
            tbl.setAlternatingRowColors(True)
            tbl.setStyleSheet(
                f"QHeaderView::section {{ background:{self._HDR_COLOR}; "
                f"color:{self._HDR_TEXT}; font-weight:600; padding:4px; "
                f"border:none; border-right:1px solid #546E7A; }}"
                f"QTableWidget {{ alternate-background-color:{self._ROW_EVEN}; "
                f"gridline-color:{self._BORDER}; }}"
            )
            for r, row in enumerate(rows):
                for c, val in enumerate(row):
                    it = QtWidgets.QTableWidgetItem(str(val))
                    it.setFlags(it.flags() & ~QtCore.Qt.ItemIsEditable)
                    align = (QtCore.Qt.AlignLeft | QtCore.Qt.AlignVCenter) if c == 0                             else (QtCore.Qt.AlignRight | QtCore.Qt.AlignVCenter)
                    it.setTextAlignment(align)
                    tbl.setItem(r, c, it)

            tbl.resizeColumnsToContents()
            if col_stretch:
                for i, s in enumerate(col_stretch):
                    if s and i < tbl.columnCount():
                        tbl.horizontalHeader().setSectionResizeMode(
                            i, QtWidgets.QHeaderView.Stretch)
            # Ajustar altura para no necesitar scroll interno
            h = tbl.horizontalHeader().height() + 4
            for r in range(tbl.rowCount()):
                h += tbl.rowHeight(r)
            tbl.setFixedHeight(min(h + 4, 400))
            parent_layout.addWidget(tbl)

        # ── Helper: par clave-valor como QFormLayout ──────────────────────
        def _form(parent_layout, pairs: list):
            form = QtWidgets.QFormLayout()
            form.setLabelAlignment(QtCore.Qt.AlignRight)
            form.setHorizontalSpacing(12)
            for label, value in pairs:
                lbl_w = QtWidgets.QLabel(f"<b>{label}:</b>")
                val_w = QtWidgets.QLabel(str(value))
                val_w.setWordWrap(True)
                form.addRow(lbl_w, val_w)
            parent_layout.addLayout(form)

        # ════════════════════════════════════════════════════════════════
        # SECCIÓN 1: Condiciones generales
        # ════════════════════════════════════════════════════════════════
        _, gl1 = _group("Condiciones generales")

        date_raw = meta.get("date", "")
        try:
            from datetime import datetime, timezone
            dt = datetime.fromisoformat(date_raw).astimezone()
            date_str = dt.strftime("%d/%m/%Y  %H:%M:%S (hora local)")
        except Exception:
            date_str = date_raw or "—"

        modo_str = "Restricción de dosis" if meta.get("mode") == "constraints" else "Tiempo fijo"
        t     = meta.get("time", 0)
        t_err = meta.get("time_err", 0)
        spnd  = meta.get("spnd", 0)
        spnd_err = meta.get("spnd_abs_err", 0)
        corr  = meta.get("Corr", None)

        pairs1 = [
            ("Fecha / hora", date_str),
            ("Tipo de dosis", dose_mode_text),
            ("Modo", modo_str),
            ("Protocolo de boro", _resolve_boro_protocol_name(
                meta.get("boro_protocol_name", "Manual"))),
            ("Origen del boro", format_boro_origin(meta)),
            ("Tiempo de irradiación",
             format_time(t, t_err, "s")),
            ("Flujo neutrónico (SPND)",
             format_scientific_value_uncertainty(spnd, spnd_err, "n/cm²·s")),
        ]
        if corr is not None:
            pairs1.append(("Factor de corrección (Corr)", f"{corr:.6g}"))
        if meta.get("dose_for_limits"):
            pairs1.append(("Dosis usada para restricciones",
                           "Biológica" if meta["dose_for_limits"] == "bio"
                           else "Física"))
        _form(gl1, pairs1)

        # ════════════════════════════════════════════════════════════════
        # SECCIÓN 2: Boro / CBE / RBE por órgano
        # ════════════════════════════════════════════════════════════════
        _, gl2 = _group("Concentración de boro y factores biológicos por órgano")

        B    = list(params.get("B",   []))
        Berr = list(params.get("B_err", []))
        CBE  = list(params.get("CBE", []))
        RBE  = list(params.get("RBE", []))

        if use_bio and not is_isoe:
            hdrs2 = ["Órgano", "B [ppm]", "σ_B [ppm]", "CBE", "RBE"]
        else:
            hdrs2 = ["Órgano", "B [ppm]", "σ_B [ppm]"]

        rows2 = []
        for j, name in enumerate(ORG_ORDER):
            b   = f"{B[j]:.4g}"   if j < len(B)    else "—"
            be  = f"{Berr[j]:.4g}" if j < len(Berr) else "—"
            cbe = f"{CBE[j]:.4g}" if j < len(CBE)  else "—"
            rbe = f"{RBE[j]:.4g}" if j < len(RBE)  else "—"
            row = [name, b, be]
            if use_bio and not is_isoe:
                row += [cbe, rbe]
            rows2.append(row)

        _table(gl2, hdrs2, rows2)

        # ════════════════════════════════════════════════════════════════
        # SECCIÓN 3: Constraints usadas (solo si modo constraints)
        # ════════════════════════════════════════════════════════════════
        constraints_used = meta.get("constraints_used") or []
        if meta.get("mode") == "constraints" and constraints_used:
            _, gl3 = _group(f"Restricciones activas — unidad: {unit}")
            hdrs3 = ["Órgano", "Tipo", f"Límite ({unit})"]
            rows3 = []
            for c in constraints_used:
                lim = c.get("limit_value", "")
                try:    lim_txt = f"{float(lim):.4g}"
                except: lim_txt = str(lim)
                rows3.append([
                    str(c.get("org", "?")),
                    str(c.get("display") or c.get("metric") or "?"),
                    lim_txt,
                ])
            _table(gl3, hdrs3, rows3, col_stretch=[True, True, False])

            # Constraint limitante
            ch = meta.get("chosen_constraint")
            if ch:
                ach = ch.get("achieved_value")
                ach_txt = f"{float(ach):.4g} {unit}" if ach is not None else "—"
                note = QtWidgets.QLabel(
                    f"<b>Restricción limitante:</b> {ch.get('org','?')} — "
                    f"{ch.get('type','?')}  |  "
                    f"Límite: {ch.get('limit_value','?')} {unit}  |  "
                    f"Logrado: {ach_txt}"
                )
                note.setWordWrap(True)
                gl3.addWidget(note)

        # ════════════════════════════════════════════════════════════════
        # SECCIÓN 4: SPND / Corrientes (solo si se calculó por corrientes)
        # ════════════════════════════════════════════════════════════════
        snap = meta.get("spnd_from_currents")
        if snap:
            _, gl4 = _group("Medición de corrientes SPND")
            hdrs4 = ["Detector", "Usar", "I medida (pA)", "± I (pA)"]
            rows4 = []
            for row in snap:
                rows4.append([
                    str(row.get("name", "?")),
                    "✓" if row.get("use") else "—",
                    f"{row.get('current', 0):.4g}",
                    f"{row.get('current_err', 0):.4g}",
                ])
            _table(gl4, hdrs4, rows4)

            # Flujo resultante
            flux     = meta.get("spnd", 0)
            flux_err = meta.get("spnd_abs_err", 0)
            gl4.addWidget(QtWidgets.QLabel(
                f"<b>Flujo resultante:</b> "
                f"{format_scientific_value_uncertainty(flux, flux_err, 'n/cm²·s')}"
            ))

        # ════════════════════════════════════════════════════════════════
        # SECCIÓN 5: Parámetros IsoE (solo si modo IsoE)
        # ════════════════════════════════════════════════════════════════
        if is_isoe:
            iso_params = meta.get("isoe_params_dict") or {}
            t0_map     = meta.get("isoe_t0_map") or {}
            preset     = meta.get("isoe_preset_name", "Manual")

            _, gl5 = _group(f"Parámetros del modelo IsoE (MLQ)  —  preset: {preset}")

            if iso_params:
                # Tabla α/β organizada por componente
                comp_labels = {
                    "R":   "Referencia fotónica",
                    "B":   "Boro",
                    "Th":  "Neutrones Térmicos",
                    "Fn":  "Neutrones Rápidos",
                    "G":   "Gamma del haz",
                }
                hdrs5 = ["Componente", "α [Gy⁻¹]", "β [Gy⁻²]", "t₀ reparación [s]"]
                rows5 = []
                for key, label in comp_labels.items():
                    a_val = iso_params.get(f"a{key}", "—")
                    b_val = iso_params.get(f"b{key}", "—")
                    t0    = t0_map.get(
                        {"R":"R","B":"Boro","Th":"Thn","Fn":"Fstn","G":"Gamma"}.get(key,""),
                        "—"
                    )
                    try:    a_txt = f"{float(a_val):.5g}"
                    except: a_txt = str(a_val)
                    try:    b_txt = f"{float(b_val):.5g}"
                    except: b_txt = str(b_val)
                    try:    t0_txt = f"{float(t0):.4g}"
                    except: t0_txt = str(t0)
                    rows5.append([label, a_txt, b_txt, t0_txt])

                # Fila extra: GR
                gr = iso_params.get("GR", "—")
                try:    gr_txt = f"{float(gr):.5g}"
                except: gr_txt = str(gr)
                rows5.append(["G_R (factor L-C de referencia)", gr_txt, "—", "—"])

                _table(gl5, hdrs5, rows5)

            note5 = QtWidgets.QLabel(
                "Los parámetros α y β son los del modelo Lineal-Cuadrático (LQ) "
                "por componente de dosis. t₀ es el tiempo de reparación "
                "subcelular (factor Lea-Catcheside)."
            )
            note5.setWordWrap(True)
            note5.setStyleSheet("color: #546E7A; font-size: 9pt;")
            gl5.addWidget(note5)

        # ── Botón cerrar ─────────────────────────────────────────────────
        vlay.addStretch()
        hb = QtWidgets.QHBoxLayout()
        hb.addStretch()
        btn_close = QtWidgets.QPushButton("Cerrar")
        btn_close.clicked.connect(self.accept)
        hb.addWidget(btn_close)
        root.addLayout(hb)

# ------------------- Diálogo Componentes de Dosis -------------------
class ComponentDosesDialog(QtWidgets.QDialog):
    """
    Muestra una tabla con las componentes de dosis por órgano:
    Boro, Neutrones Rápidos (Fstn), Neutrones Térmicos (Thn) y Gamma.
    Para cada componente se muestran Dosis Mínima, Dosis Media y Dosis Máxima.
    Los datos provienen de report["CompDoses"] que ya están multiplicados por el tiempo.

    La tabla usa dos filas de encabezado incrustadas (setSpan):
      Fila 0 → título del grupo (ej. "Dosis Boro")   [span 3 columnas]
      Fila 1 → subcolumnas: Dosis Mínima | Dosis Media | Dosis Máxima
    """
    def __init__(self, parent, report, dose_mode_text, use_bio=False, is_isoe=False):
        super().__init__(parent)
        self.setWindowTitle("Componentes de Dosis por Órgano")
        self.resize(1300, 500)
        self.report = report
        self.dose_mode_text = dose_mode_text
        self.use_bio = bool(use_bio)
        self.is_isoe = bool(is_isoe)

        layout = QtWidgets.QVBoxLayout(self)

        unit = dose_axis_unit(use_bio, is_isoe)
        note = QtWidgets.QLabel(
            f"<b>Unidades:</b> {unit} &nbsp;&nbsp; "
            "<b>Nota:</b> Componentes de dosis física multiplicadas por el tiempo de irradiación. "
            "Los factores RBE/CBE <u>no</u> están aplicados en esta tabla."
        )
        note.setWordWrap(True)
        layout.addWidget(note)

        comp_data = report.get("CompDoses", {})

        # Componentes: (clave en CompDoses, título del grupo)
        COMPS = [
            ("Boro",  "Dosis Boro"),
            ("Fstn",  "Dosis Neutrones Rápidos"),
            ("Thn",   "Dosis Neutrones Térmicos"),
            ("Gamma", "Dosis Gamma"),
        ]
        SUB_COLS  = ["Dosis Promedio", "Dosis Mínima", "Dosis Máxima"]
        organs    = sorted(comp_data.keys())
        n_sub     = len(SUB_COLS)          # 3
        n_comps   = len(COMPS)             # 4
        # Columnas: Órgano + 4 grupos×3 subcolumnas + Total×3 = 1+12+3 = 16
        n_cols    = 1 + n_comps * n_sub + n_sub
        N_HDR     = 2                       # filas de encabezado embebidas
        n_rows    = N_HDR + len(organs)

        # ── Colores y fuentes ─────────────────────────────────────────────
        HDR1_BG = QtGui.QColor("#455A64")   # fila 0: título de grupo
        HDR1_FG = QtGui.QColor("#FFFFFF")
        HDR2_BG = QtGui.QColor("#607D8B")   # fila 1: subcolumnas
        HDR2_FG = QtGui.QColor("#FFFFFF")
        EVEN_BG = QtGui.QColor("#F5F7FA")
        ODD_BG  = QtGui.QColor("#FFFFFF")

        bold_font = QtGui.QFont()
        bold_font.setBold(True)

        def _hdr(text, bg, fg):
            """Celda de encabezado embebida."""
            it = QtWidgets.QTableWidgetItem(text)
            it.setFlags(QtCore.Qt.ItemIsEnabled)
            it.setBackground(bg)
            it.setForeground(fg)
            it.setFont(bold_font)
            it.setTextAlignment(QtCore.Qt.AlignCenter)
            return it

        def _data(text, right=True, bg=None):
            """Celda de datos."""
            it = QtWidgets.QTableWidgetItem(str(text))
            it.setFlags(QtCore.Qt.ItemIsEnabled | QtCore.Qt.ItemIsSelectable)
            align = (QtCore.Qt.AlignRight | QtCore.Qt.AlignVCenter) if right \
                    else (QtCore.Qt.AlignLeft  | QtCore.Qt.AlignVCenter)
            it.setTextAlignment(align)
            if bg is not None:
                it.setBackground(bg)
            return it

        # ── Crear tabla ───────────────────────────────────────────────────
        tbl = QtWidgets.QTableWidget(n_rows, n_cols)
        tbl.setAlternatingRowColors(False)   # gestionamos color manualmente
        tbl.setShowGrid(True)
        tbl.verticalHeader().setVisible(False)
        tbl.horizontalHeader().setVisible(False)   # usamos filas embebidas
        tbl.setEditTriggers(QtWidgets.QAbstractItemView.NoEditTriggers)
        tbl.setSelectionMode(QtWidgets.QAbstractItemView.SingleSelection)

        # ── Fila 0: títulos de grupo ──────────────────────────────────────
        # Columna "Órgano" abarca las dos filas de encabezado
        tbl.setItem(0, 0, _hdr("Órgano", HDR1_BG, HDR1_FG))
        tbl.setSpan(0, 0, N_HDR, 1)

        for i, (_, comp_label) in enumerate(COMPS):
            col_start = 1 + i * n_sub
            tbl.setItem(0, col_start, _hdr(comp_label, HDR1_BG, HDR1_FG))
            tbl.setSpan(0, col_start, 1, n_sub)

        total_col_start = 1 + n_comps * n_sub
        tbl.setItem(0, total_col_start, _hdr("Dosis Total", HDR1_BG, HDR1_FG))
        tbl.setSpan(0, total_col_start, 1, n_sub)

        # ── Fila 1: subcolumnas ───────────────────────────────────────────
        for group in range(n_comps + 1):   # 4 componentes + 1 total
            for j, sub_label in enumerate(SUB_COLS):
                col = 1 + group * n_sub + j
                tbl.setItem(1, col, _hdr(sub_label, HDR2_BG, HDR2_FG))

        # Altura de las filas de encabezado
        tbl.setRowHeight(0, 30)
        tbl.setRowHeight(1, 26)

        # ── Filas de datos ────────────────────────────────────────────────
        for ridx, organ in enumerate(organs):
            row = N_HDR + ridx
            bg  = EVEN_BG if ridx % 2 == 0 else ODD_BG
            cd  = comp_data[organ]

            arrays = {}
            for comp_key, _ in COMPS:
                arr = np.array(cd.get(comp_key, []), dtype=float)
                arrays[comp_key] = arr[np.isfinite(arr)]

            # Total físico como suma de componentes
            valid = [arrays[k] for k in arrays if arrays[k].size > 0]
            if valid:
                # Si todos tienen el mismo tamaño, suma directa; si no, suma parcial
                ref_size = valid[0].size
                total = np.zeros(ref_size)
                for arr in valid:
                    n = min(arr.size, ref_size)
                    total[:n] += arr[:n]
            else:
                total = np.array([0.0])

            # Columna Órgano
            tbl.setItem(row, 0, _data(organ, right=False, bg=bg))

            col = 1
            for comp_key, _ in COMPS:
                arr = arrays[comp_key]
                if arr.size > 0:
                    dmin  = float(np.nanmin(arr))
                    dmean = float(np.nanmean(arr))
                    dmax  = float(np.nanmax(arr))
                else:
                    dmin = dmean = dmax = 0.0
                for val in (dmean, dmin, dmax):   # orden: Promedio, Mínima, Máxima
                    tbl.setItem(row, col, _data(f"{val:.4f}", bg=bg))
                    col += 1

            # Total
            if total.size > 0:
                t_min  = float(np.nanmin(total))
                t_mean = float(np.nanmean(total))
                t_max  = float(np.nanmax(total))
            else:
                t_min = t_mean = t_max = 0.0
            for val in (t_mean, t_min, t_max):   # orden: Promedio, Mínima, Máxima
                tbl.setItem(row, col, _data(f"{val:.4f}", bg=bg))
                col += 1

        tbl.resizeColumnsToContents()
        _install_copy_shortcut(tbl)
        layout.addWidget(tbl)

        btn_close = QtWidgets.QPushButton("Cerrar")
        btn_close.clicked.connect(self.accept)
        h = QtWidgets.QHBoxLayout()
        h.addStretch()
        h.addWidget(btn_close)
        layout.addLayout(h)


# ------------------- Diálogo Resultados -------------------
class ResultsDialog(QtWidgets.QDialog):
    def __init__(self, parent, report, dsum_selected, deq, dose_mode_text,
                 use_bio=False, is_isoe=False, viz_data=None,
                 isoe_params_by_organ=None):
        super().__init__(parent)
        self.setWindowTitle("Resultados")
        self.resize(1280, 600)
        self.setMinimumSize(720, 480)
        self.report = report
        self.dsum = dsum_selected
        self.deq = deq
        self.dose_mode_text = dose_mode_text
        self.use_bio = bool(use_bio)
        self.is_isoe = bool(is_isoe)
        self.viz_data = viz_data
        self.isoe_params_by_organ = isoe_params_by_organ or {}

        layout = QtWidgets.QVBoxLayout(self)

        # Unidad de dosis según modo activo (usada en todo el diálogo y el PDF)
        unit = dose_axis_unit(use_bio, is_isoe)

        # encabezado
        meta = report.get("meta", {})
        self._meta = meta   # guardado para re-formatear tiempo si cambia el modo
        self._unit = unit
        self._time_fmt = "s"   # "s" o "ms" — controlado por el combo de abajo

        hdr = QtWidgets.QLabel()
        hdr.setTextFormat(QtCore.Qt.RichText)
        self._hdr = hdr   # guardado para actualizar al cambiar formato de tiempo

        def _build_hdr():
            t   = meta.get("time", 0)
            t_e = meta.get("time_err", 0)
            s = f"<b>Tipo de cálculo:</b> {dose_mode_text} &nbsp;&nbsp; "
            s += f"<b>Tiempo:</b> {format_time(t, t_e, self._time_fmt)}"
            if meta.get("mode", "time") == "constraints":
                ch = meta.get("chosen_constraint")
                if ch:
                    ach = ch.get("achieved_value", 0.0)
                    s += (f" &nbsp;&nbsp; <b>Restricción limitante:</b> "
                          f"{ch.get('org','?')} - {ch.get('type','?')} "
                          f"(límite {ch.get('limit_value','?')} {unit}, "
                          f"logrado {float(ach):.3f} {unit})")
            return s

        self._build_hdr = _build_hdr
        hdr.setText(_build_hdr())
        layout.addWidget(hdr)

        # parámetros (valores seleccionables con el mouse)
        pbox = QtWidgets.QGroupBox("Parámetros")
        form = QtWidgets.QFormLayout(pbox)
        form.addRow("Protocolo de boro:",
                    _selectable_label(_resolve_boro_protocol_name(meta.get("boro_protocol_name", "Manual"))))
        form.addRow("Origen del boro:",
                    _selectable_label(format_boro_origin(meta)))
        form.addRow("Flujo neutrónico (n/cm\u00b2·s):",
                    _selectable_label(format_scientific_value_uncertainty(
                        meta.get("spnd", 0), meta.get("spnd_abs_err", 0), "n/cm²·s")))

        # Tiempo con selector de formato (segundos / minutos+segundos)
        self._lbl_tiempo = _selectable_label(
            format_time(meta.get("time", 0), meta.get("time_err", 0), self._time_fmt))
        cmb_time_fmt = QtWidgets.QComboBox()
        cmb_time_fmt.addItem("en segundos", "s")
        cmb_time_fmt.addItem("en min + seg", "ms")
        cmb_time_fmt.setFixedWidth(120)
        time_row = QtWidgets.QHBoxLayout()
        time_row.addWidget(self._lbl_tiempo)
        time_row.addStretch()
        time_row.addWidget(QtWidgets.QLabel("Mostrar"))
        time_row.addWidget(cmb_time_fmt)
        time_widget = QtWidgets.QWidget()
        time_widget.setLayout(time_row)
        form.addRow("Tiempo:", time_widget)

        def _on_time_fmt_changed():
            self._time_fmt = cmb_time_fmt.currentData()
            t   = meta.get("time", 0)
            t_e = meta.get("time_err", 0)
            self._lbl_tiempo.setText(format_time(t, t_e, self._time_fmt))
            self._hdr.setText(self._build_hdr())

        cmb_time_fmt.currentIndexChanged.connect(_on_time_fmt_changed)
        layout.addWidget(pbox)

        # ── Tabla de métricas (panel izquierdo) ──────────────────────────────
        # Voxmap para calcular métricas adicionales (Dx%) — guardado como
        # atributo para usarlo al togglear la columna extra al maximizar.
        self._voxmap_for_metric = self.report.get(
            "IsoVoxel" if self.is_isoe else ("BioVoxel" if self.use_bio else "PhysVoxel"), {}
        )

        # Columnas base (siempre visibles): Órgano | Promedio ± | Mínima ± | Máxima ±
        # Columnas extendidas (solo al maximizar): D95 ± | D5 ± | Dx% (col extra)
        self._cols_base     = ["Órgano", "Dosis Promedio", "±", "Dosis Mínima", "±", "Dosis Máxima", "±"]
        self._cols_extended = ["D95", "±", "D5", "±"]
        self._col_dx_label  = "D10%"   # encabezado dinámico de la columna extra

        self.table = QtWidgets.QTableWidget()
        self.table.setAlternatingRowColors(True)
        self.table.setShowGrid(True)
        self.table.verticalHeader().setVisible(False)
        keys = sorted(self.dsum.keys())
        self._table_keys = keys
        self.table.setRowCount(len(keys))
        self._build_table_columns(extended=False)

        _install_copy_shortcut(self.table)

        # Spinbox para la columna Dx% (visible solo cuando la tabla está maximizada)
        self.spin_pct = QtWidgets.QSpinBox()
        self.spin_pct.setRange(1, 99)
        self.spin_pct.setValue(10)
        self.spin_pct.setSuffix(" %")
        self.spin_pct.setToolTip("Porcentaje de volumen — D10% = dosis que recibe el 10% más caliente del volumen")
        self.spin_pct.valueChanged.connect(self._refresh_dx_column)
        self._spin_pct_widget = QtWidgets.QWidget()
        spin_row = QtWidgets.QHBoxLayout(self._spin_pct_widget)
        spin_row.setContentsMargins(0, 0, 0, 0)
        spin_row.addWidget(QtWidgets.QLabel("Mostrar columna D"))
        spin_row.addWidget(self.spin_pct)
        spin_row.addWidget(QtWidgets.QLabel("del vol."))
        spin_row.addStretch()
        self._spin_pct_widget.setVisible(False)  # oculto hasta que se maximice la tabla

        table_panel = QtWidgets.QWidget()
        table_panel_layout = QtWidgets.QVBoxLayout(table_panel)
        table_panel_layout.setContentsMargins(0, 0, 0, 0)
        table_header = QtWidgets.QHBoxLayout()
        lbl_metrics_title = QtWidgets.QLabel("<b>Métricas dosimétricas</b>")
        lbl_metrics_title.setAlignment(QtCore.Qt.AlignCenter)
        table_header.addStretch()
        table_header.addWidget(lbl_metrics_title)
        table_header.addStretch()
        self.btn_max_table = QtWidgets.QToolButton()
        self.btn_max_table.setText("Maximizar ↔")
        self.btn_max_table.setCheckable(True)
        self.btn_max_table.setToolTip("Expande la tabla y oculta temporalmente el gráfico")
        table_header.addWidget(self.btn_max_table)
        table_panel_layout.addLayout(table_header)
        table_panel_layout.addWidget(self.table, 1)         # stretch=1: tabla ocupa todo el espacio vertical
        table_panel_layout.addWidget(self._spin_pct_widget) # fijo abajo, solo visible al maximizar

        # ── DVH (panel derecho) ───────────────────────────────────────────────
        dvh_panel = QtWidgets.QWidget()
        dvh_panel_layout = QtWidgets.QVBoxLayout(dvh_panel)
        dvh_panel_layout.setContentsMargins(0, 0, 0, 0)
        dvh_header = QtWidgets.QHBoxLayout()
        lbl_dvh_title = QtWidgets.QLabel("<b>Histograma dosis-volumen (DVH)</b>")
        lbl_dvh_title.setAlignment(QtCore.Qt.AlignCenter)
        dvh_header.addStretch()
        dvh_header.addWidget(lbl_dvh_title)
        dvh_header.addStretch()
        self.btn_max_dvh = QtWidgets.QToolButton()
        self.btn_max_dvh.setText("Maximizar ↔")
        self.btn_max_dvh.setCheckable(True)
        self.btn_max_dvh.setToolTip("Expande el gráfico y oculta temporalmente la tabla")
        dvh_header.addWidget(self.btn_max_dvh)
        dvh_panel_layout.addLayout(dvh_header)

        self.canvas = MplCanvas(self, width=6, height=4.5)
        self.dvh_toolbar = NavigationToolbar2QT(self.canvas, self)
        dvh_panel_layout.addWidget(self.dvh_toolbar)

        self.canvas.ax.clear()
        xlab = f"Dosis ({dose_axis_unit(self.use_bio, self.is_isoe)})"
        title = "DVH — Dosis Isoefectiva (Photon IsoE)" if self.is_isoe else (
                "DVH — Dosis Equivalente pesada por RBE" if self.use_bio else "DVH — Dosis Física")
        self.canvas.ax.set_xlabel(xlab)
        self.canvas.ax.set_ylabel("Volumen (%)")
        self.canvas.ax.set_title(title)
        voxmap = self._voxmap_for_metric
        _keys = list(voxmap.keys())
        _color_map = {k: ORGAN_COLORS[i % len(ORGAN_COLORS)] for i, k in enumerate(_keys)}
        for key, vox in voxmap.items():
            arr = np.array(vox, dtype=float)
            if arr.size == 0: continue
            s, vol = build_dvh(arr)
            s, vol = dvh_extend_to_zero(s, vol)
            if s.size > 0:
                self.canvas.ax.plot(s, vol, label=key, color=_color_map[key], picker=5)
        self.canvas.ax.set_xlim(left=0)
        self.canvas.ax.set_ylim(0, 105)
        self.canvas.ax.grid(True, which='major', color='#E0E0E0', linewidth=0.8, alpha=0.9)
        self.canvas.ax.grid(True, which='minor', color='#EEEEEE', linewidth=0.4, alpha=0.7)
        self.canvas.ax.minorticks_on()
        self.canvas.ax.set_axisbelow(True)
        self.canvas.ax.legend(loc="best", fontsize=8, framealpha=0.85, edgecolor='#CCCCCC')
        self.canvas.fig.tight_layout(pad=1.5)
        self.canvas.draw()
        dvh_panel_layout.addWidget(self.canvas, 1)   # stretch=1 para que ocupe todo el espacio vertical

        # Label de coordenadas — compacto, sin texto explicativo para no robar espacio
        self.lbl_coords = QtWidgets.QLabel("")
        self.lbl_coords.setStyleSheet("color: #546E7A; font-size: 11px;")
        self.lbl_coords.setFixedHeight(16)
        dvh_panel_layout.addWidget(self.lbl_coords)
        self.canvas.mpl_connect("button_press_event", self._on_dvh_click)

        # ── Splitter: tabla a la izquierda, DVH a la derecha ─────────────────
        self.results_splitter = QtWidgets.QSplitter(QtCore.Qt.Horizontal)
        self.results_splitter.addWidget(table_panel)
        self.results_splitter.addWidget(dvh_panel)
        self.results_splitter.setStretchFactor(0, 1)
        self.results_splitter.setStretchFactor(1, 1)
        self.results_splitter.setSizes([1, 1])
        layout.addWidget(self.results_splitter, 1)

        self._table_panel = table_panel
        self._dvh_panel = dvh_panel
        self.btn_max_table.toggled.connect(self._on_toggle_max_table)
        self.btn_max_dvh.toggled.connect(self._on_toggle_max_dvh)

        # botones
        h = QtWidgets.QHBoxLayout()

        # Botón "Exportar resultados" con menú desplegable (PDF / Excel / Ambos)
        self.btn_export = QtWidgets.QPushButton("Exportar resultados ▾")
        self.btn_export.setStyleSheet("QPushButton::menu-indicator { image: none; }")
        self.btn_export.setToolTip(
            "PDF: reporte dosimétrico completo.\n"
            "Excel: tablas de métricas y componentes + parámetros del cálculo.\n"
            "PDF + Excel: genera ambos archivos."
        )
        _export_menu = QtWidgets.QMenu(self.btn_export)
        self._act_pdf   = _export_menu.addAction("Exportar PDF")
        self._act_xlsx  = _export_menu.addAction("Exportar Excel (.xlsx)")
        self._act_both  = _export_menu.addAction("Exportar PDF + Excel")
        self.btn_export.setMenu(_export_menu)

        self.btn_comp   = QtWidgets.QPushButton("Componentes de Dosis")
        self.btn_comp.setToolTip("Muestra la tabla de componentes: Boro, Neutrones Rápidos, Térmicos y Gamma por órgano")
        self.btn_params = QtWidgets.QPushButton("Parámetros del cálculo")
        self.btn_params.setToolTip("Muestra todos los parámetros usados: boro, CBE/RBE, SPND, restricciones, IsoE")
        self.btn_viz    = QtWidgets.QPushButton("Visualizar Dosis")
        self.btn_viz.setToolTip(
            "Abre el visor 2D de mapas de dosis por órgano superpuestos a la segmentación.\n"
            "Requiere que se hayan cargado datos espaciales (build_viz_data)."
        )
        # El botón se habilita solo si se pasó viz_data con información espacial
        _viz_available = (
            self.viz_data is not None
            and bool(self.viz_data.get("organ_valid_indices"))
        )
        self.btn_viz.setEnabled(_viz_available)
        if not _viz_available:
            self.btn_viz.setToolTip(
                "No disponible: los datos espaciales de segmentación no fueron provistos.\n"
                "Llamá a build_viz_data() y pasá el resultado como viz_data al constructor."
            )

        self.btn_radiobio = QtWidgets.QPushButton("Prob. Dosis-Efecto")
        self.btn_radiobio.setToolTip(
            "Calcula TCP, NTCP y supervivencia celular a partir de la dosis isoefectiva.\n"
            "Requiere modo IsoE activo."
        )
        _rb_ok = (RADIOBIO_OK and self.is_isoe
                  and bool(self.report.get("IsoVoxel"))
                  and bool(self.isoe_params_by_organ))
        self.btn_radiobio.setEnabled(_rb_ok)
        if not _rb_ok:
            tip = ("Módulo radiobio no instalado." if not RADIOBIO_OK
                   else "Solo disponible en modo IsoE con parámetros asignados.")
            self.btn_radiobio.setToolTip(tip)

        self.btn_close  = QtWidgets.QPushButton("Cerrar")
        h.addWidget(self.btn_export)
        h.addWidget(self.btn_comp)
        h.addWidget(self.btn_params)
        h.addWidget(self.btn_viz)
        h.addWidget(self.btn_radiobio)
        h.addStretch()
        h.addWidget(self.btn_close)
        layout.addLayout(h)

        self.btn_close.clicked.connect(self.accept)
        self._act_pdf.triggered.connect(self._export_pdf)
        self._act_xlsx.triggered.connect(self._export_xlsx)
        self._act_both.triggered.connect(self._export_both)
        self.btn_comp.clicked.connect(self._open_comp_doses)
        self.btn_params.clicked.connect(self._open_calc_params)
        self.btn_viz.clicked.connect(self._open_dose_viewer)
        self.btn_radiobio.clicked.connect(self._open_radiobio_dialog)

    # ── helpers de tabla ──────────────────────────────────────────────────────

    def _build_table_columns(self, extended: bool):
        """
        Construye (o reconstruye) las columnas de la tabla de métricas.

        extended=False → columnas base: Órgano | Promedio ± | Mínima ± | Máxima ±
        extended=True  → columnas base + D95 ± | D5 ± | Dx% (columna extra dinámica)

        Se puede llamar varias veces — resetea el contenido y vuelve a llenarlo.
        """
        def it(txt):
            item = QtWidgets.QTableWidgetItem(str(txt))
            item.setFlags(item.flags() & ~QtCore.Qt.ItemIsEditable)
            return item

        pct = self.spin_pct.value() if hasattr(self, "spin_pct") else 10
        dx_label = f"D{pct}%"

        if extended:
            cols = self._cols_base + self._cols_extended + [dx_label]
        else:
            cols = self._cols_base

        self.table.setColumnCount(len(cols))
        self.table.setHorizontalHeaderLabels(cols)

        for i, k in enumerate(self._table_keys):
            v = self.dsum[k]
            dmean_v, dmean_s = format_value_uncertainty_separate(v["Dmean"], v["Sigma_Dmean"])
            dmin_v,  dmin_s  = format_value_uncertainty_separate(v["Dmin"],  v["Sigma_Dmin"])
            dmax_v,  dmax_s  = format_value_uncertainty_separate(v["Dmax"],  v["Sigma_Dmax"])

            self.table.setItem(i, 0, it(k))
            self.table.setItem(i, 1, it(dmean_v))
            self.table.setItem(i, 2, it(dmean_s))
            self.table.setItem(i, 3, it(dmin_v))
            self.table.setItem(i, 4, it(dmin_s))
            self.table.setItem(i, 5, it(dmax_v))
            self.table.setItem(i, 6, it(dmax_s))

            if extended:
                d95_v, d95_s = format_value_uncertainty_separate(v["D95"], v["Sigma_D95"])
                d5_v,  d5_s  = format_value_uncertainty_separate(v["D5"],  v["Sigma_D5"])
                self.table.setItem(i, 7, it(d95_v))
                self.table.setItem(i, 8, it(d95_s))
                self.table.setItem(i, 9, it(d5_v))
                self.table.setItem(i, 10, it(d5_s))
                # Columna Dx%
                dx_val = achieved_value_for_constraint(self._voxmap_for_metric, k, f"D{pct}%")
                unit = dose_axis_unit(self.use_bio, self.is_isoe)
                self.table.setItem(i, 11, it(f"{dx_val:.2f} {unit}" if dx_val is not None else "—"))

        self.table.resizeColumnsToContents()

    def _refresh_dx_column(self):
        """
        Actualiza el encabezado y los valores de la columna Dx% cuando
        cambia el spinbox — solo cuando la tabla está en modo extendido.
        """
        if not self.btn_max_table.isChecked():
            return
        pct = self.spin_pct.value()
        dx_label = f"D{pct}%"
        ncols = self.table.columnCount()
        dx_col = ncols - 1   # siempre la última columna
        self.table.setHorizontalHeaderItem(dx_col, QtWidgets.QTableWidgetItem(dx_label))
        unit = dose_axis_unit(self.use_bio, self.is_isoe)
        for i, k in enumerate(self._table_keys):
            dx_val = achieved_value_for_constraint(self._voxmap_for_metric, k, f"D{pct}%")
            item = QtWidgets.QTableWidgetItem(f"{dx_val:.2f} {unit}" if dx_val is not None else "—")
            item.setFlags(item.flags() & ~QtCore.Qt.ItemIsEditable)
            self.table.setItem(i, dx_col, item)
        self.table.resizeColumnToContents(dx_col)

    # ── DVH: click con coordenadas ────────────────────────────────────────────

    def _on_dvh_click(self, event):
        """Muestra las coordenadas (dosis, volumen%) del punto donde se hizo click."""
        if event.inaxes != self.canvas.ax or event.xdata is None or event.ydata is None:
            return
        unit = dose_axis_unit(self.use_bio, self.is_isoe)
        self.lbl_coords.setText(
            f"Dosis: {event.xdata:.2f} {unit}   |   Volumen: {event.ydata:.1f} %"
        )

    # ── toggles de maximizar ──────────────────────────────────────────────────

    def _on_toggle_max_table(self, checked: bool):
        """
        Maximiza la tabla (oculta el DVH) o restaura el 50/50.
        Al maximizar: muestra columnas extendidas (D95, D5, Dx%) y el spinbox.
        Al restaurar: vuelve a columnas base y oculta el spinbox.
        Texto del botón cambia entre "Maximizar ↔" y "Restaurar ↔".
        """
        if checked:
            self.btn_max_table.setText("Restaurar ↔")
            self.btn_max_dvh.setChecked(False)
            self._dvh_panel.setVisible(False)
            total = sum(self.results_splitter.sizes()) or self.results_splitter.width()
            self.results_splitter.setSizes([total, 0])
            self._build_table_columns(extended=True)
            self._spin_pct_widget.setVisible(True)
        else:
            self.btn_max_table.setText("Maximizar ↔")
            self._dvh_panel.setVisible(True)
            self.results_splitter.setSizes([1, 1])
            self._build_table_columns(extended=False)
            self._spin_pct_widget.setVisible(False)

    def _on_toggle_max_dvh(self, checked: bool):
        """
        Maximiza el DVH (oculta la tabla) o restaura el 50/50.
        Texto del botón cambia entre "Maximizar ↔" y "Restaurar ↔".
        """
        if checked:
            self.btn_max_dvh.setText("Restaurar ↔")
            self.btn_max_table.setChecked(False)
            self._table_panel.setVisible(False)
            total = sum(self.results_splitter.sizes()) or self.results_splitter.width()
            self.results_splitter.setSizes([0, total])
        else:
            self.btn_max_dvh.setText("Maximizar ↔")
            self._table_panel.setVisible(True)
            self.results_splitter.setSizes([1, 1])

    def _open_radiobio_dialog(self):
        """Abre RadiobioDialog para análisis TCP/NTCP y supervivencia."""
        if not RADIOBIO_OK:
            QtWidgets.QMessageBox.warning(self, "Módulo no disponible",
                "El módulo physics/radiobio no está instalado.")
            return
        if not self.is_isoe or not self.report.get("IsoVoxel"):
            QtWidgets.QMessageBox.information(self, "IsoE requerido",
                "El análisis radiobiológico requiere dosis isoefectiva (modo IsoE).")
            return
        from ratmaster.ui.dialogs.radiobio import RadiobioDialog
        dlg = RadiobioDialog(
            self,
            report=self.report,
            isoe_params_by_organ=self.isoe_params_by_organ,
        )
        dlg.exec()

    def _open_dose_viewer(self):
        """Abre DoseViewerDialog con el visor 2D de mapas de dosis."""
        if self.viz_data is None or not self.viz_data.get("organ_valid_indices"):
            QtWidgets.QMessageBox.information(
                self, "Sin datos espaciales",
                "No hay datos espaciales disponibles para visualizar.\n\n"
                "Asegurate de llamar a build_viz_data() después del cálculo\n"
                "y de pasar el resultado como viz_data al constructor de ResultsDialog."
            )
            return
        dlg = DoseViewerDialog(
            self,
            viz_data=self.viz_data,
            report=self.report,
            use_bio=self.use_bio,
            is_isoe=self.is_isoe,
        )
        dlg.exec()

    def _open_comp_doses(self):
        """Abre el diálogo con la tabla de componentes de dosis por órgano."""
        if not self.report.get("CompDoses"):
            QtWidgets.QMessageBox.information(self, "Sin datos", "No hay componentes de dosis disponibles en este reporte.")
            return
        dlg = ComponentDosesDialog(self, self.report, self.dose_mode_text, self.use_bio, self.is_isoe)
        dlg.exec()

    def _open_calc_params(self):
        """Abre el diálogo con todos los parámetros usados en el cálculo."""
        dlg = CalculationParamsDialog(
            self, self.report, self.dose_mode_text,
            use_bio=self.use_bio, is_isoe=self.is_isoe
        )
        dlg.exec()

    def _format_chosen_constraint(self, report):
        try:
            ch = (report.get("meta", {}) or {}).get("chosen_constraint", {}) or {}
            org = ch.get("org") or ch.get("organ") or "?"
            typ = ch.get("type") or ch.get("metric") or "?"
            _unit = dose_axis_unit(self.use_bio, self.is_isoe)
            lim = ch.get("limit_value") if ch.get("limit_value") is not None else ch.get("limit") or "?"
            ach = ch.get("achieved_value")
            if ach is None and isinstance(self.dsum, dict) and org in self.dsum:
                # Try to infer from metrics dict (Dmax/Dmean). Vx% omitted if custom.
                dm = self.dsum[org]
                key = {"Dmax":"Dmax", "Dmean":"Dmean"}.get(typ, None)
                if key and key in dm:
                    ach = dm.get(key)
            if ach is None:
                return f" &nbsp;&nbsp; <b>Restricción limitante:</b> {org} - {typ} (límite {lim})"
            else:
                try:
                    ach_val = float(ach)
                    return f" &nbsp;&nbsp; <b>Restricción limitante:</b> {org} - {typ} (límite {lim}, logrado {ach_val:.3f})"
                except Exception:
                    return f" &nbsp;&nbsp; <b>Restricción limitante:</b> {org} - {typ} (límite {lim}, logrado {ach})"
        except Exception:
            return ""
    def _export_both(self):
        """Exporta PDF y Excel en secuencia."""
        self._export_pdf()
        self._export_xlsx()

    def _export_xlsx(self):
        """Exporta un archivo Excel (.xlsx) con métricas, componentes y parámetros."""
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.utils import get_column_letter
        except ImportError:
            QtWidgets.QMessageBox.warning(
                self, "Excel no disponible",
                "Instalá openpyxl para exportar Excel:\n  pip install openpyxl"
            )
            return

        sugerido = Path.home() / (
            "resultados_ratmaster_" + datetime.now().strftime("%Y%m%dT%H%M%S") + ".xlsx"
        )
        fname, _ = QtWidgets.QFileDialog.getSaveFileName(
            self, "Guardar Excel", str(sugerido), "Excel (*.xlsx)"
        )
        if not fname:
            return

        wb = openpyxl.Workbook()

        # ── Estilos ──────────────────────────────────────────────────────────
        hdr_fill  = PatternFill("solid", fgColor="455A64")
        hdr_font  = Font(bold=True, color="FFFFFF")
        hdr2_fill = PatternFill("solid", fgColor="607D8B")
        hdr2_font = Font(bold=True, color="FFFFFF")
        meta_fill = PatternFill("solid", fgColor="F0F0F0")
        thin      = Side(style="thin", color="CFD8DC")
        border    = Border(left=thin, right=thin, top=thin, bottom=thin)
        center    = Alignment(horizontal="center", vertical="center", wrap_text=True)
        right_al  = Alignment(horizontal="right", vertical="center")
        left_al   = Alignment(horizontal="left",  vertical="center")

        def _hdr_cell(ws, row, col, text, fill=None, font=None):
            c = ws.cell(row=row, column=col, value=text)
            c.fill = fill or hdr_fill
            c.font = font or hdr_font
            c.alignment = center
            c.border = border
            return c

        def _data_cell(ws, row, col, value, align=None):
            c = ws.cell(row=row, column=col, value=value)
            c.border = border
            c.alignment = align or right_al
            return c

        meta  = self.report.get("meta", {}) or {}
        unit  = dose_axis_unit(self.use_bio, self.is_isoe)

        # ═══════════════════════════════════════════════════════════════════
        # HOJA 1: Parámetros del cálculo
        # ═══════════════════════════════════════════════════════════════════
        ws_meta = wb.active
        ws_meta.title = "Parámetros"

        ws_meta.column_dimensions["A"].width = 32
        ws_meta.column_dimensions["B"].width = 42

        _hdr_cell(ws_meta, 1, 1, "Parámetros del cálculo")
        ws_meta.merge_cells("A1:B1")

        # Fecha
        date_raw = meta.get("date", "")
        try:
            from datetime import timezone as _tz
            _dt = datetime.fromisoformat(date_raw).astimezone()
            date_str = _dt.strftime("%d/%m/%Y  %H:%M:%S")
        except Exception:
            date_str = date_raw or "—"

        modo_str = "Restricciones de dosis" if meta.get("mode") == "constraints" else "Tiempo fijo"

        pairs = [
            ("Fecha / hora",              date_str),
            ("Tipo de dosis",             self.dose_mode_text),
            ("Modo de cálculo",           modo_str),
            ("Protocolo de boro",         _resolve_boro_protocol_name(meta.get("boro_protocol_name", "Manual"))),
            ("Origen del boro",           format_boro_origin(meta)),
            ("Tiempo de irradiación",     format_time(meta.get("time", 0), meta.get("time_err", 0), "s")),
            ("Flujo neutrónico (n/cm²·s)",format_scientific_value_uncertainty(meta.get("spnd", 0), meta.get("spnd_abs_err", 0), "n/cm²·s")),
        ]

        # Constraint limitante
        if meta.get("mode") == "constraints" and meta.get("chosen_constraint"):
            ch = meta["chosen_constraint"]
            ach = ch.get("achieved_value")
            ach_txt = f"{float(ach):.3f} {unit}" if ach is not None else "—"
            pairs.append(("Restricción limitante",
                          f"{ch.get('org','?')} — {ch.get('type','?')} "
                          f"(límite {ch.get('limit_value','?')} {unit}, logrado {ach_txt})"))

        for ridx, (label, value) in enumerate(pairs, start=2):
            c_lbl = ws_meta.cell(row=ridx, column=1, value=label)
            c_lbl.fill = meta_fill
            c_lbl.font = Font(bold=True)
            c_lbl.border = border
            c_lbl.alignment = left_al
            c_val = ws_meta.cell(row=ridx, column=2, value=value)
            c_val.border = border
            c_val.alignment = left_al

        # ═══════════════════════════════════════════════════════════════════
        # HOJA 2: Métricas de dosis
        # ═══════════════════════════════════════════════════════════════════
        ws_res = wb.create_sheet("Métricas de dosis")

        col_widths = [18, 14, 10, 14, 10, 14, 10, 12, 10, 12, 10]
        for ci, w in enumerate(col_widths, start=1):
            ws_res.column_dimensions[get_column_letter(ci)].width = w

        head_mode = (
            "Dosis Isoefectiva — Gy(IsoE)" if self.is_isoe
            else ("Dosis Equivalente RBE — Gy(RBE)" if self.use_bio else "Dosis Física — Gy")
        )
        _hdr_cell(ws_res, 1, 1, head_mode)
        ws_res.merge_cells(f"A1:{get_column_letter(len(col_widths))}1")

        hdr_cols = [
            "Órgano",
            f"Dosis Promedio ({unit})", "± σ",
            f"Dosis Mínima ({unit})", "± σ",
            f"Dosis Máxima ({unit})", "± σ",
            "D95", "± σ",
            "D5",  "± σ",
        ]
        for ci, h_txt in enumerate(hdr_cols, start=1):
            _hdr_cell(ws_res, 2, ci, h_txt, fill=hdr2_fill, font=hdr2_font)

        for ridx, key in enumerate(sorted(self.dsum.keys()), start=3):
            v = self.dsum[key]
            fill_row = PatternFill("solid", fgColor="F5F7FA" if ridx % 2 == 0 else "FFFFFF")
            vals = [
                key,
                round(v["Dmean"], 4),       round(v["Sigma_Dmean"], 4),
                round(v["Dmin"],  4),       round(v["Sigma_Dmin"],  4),
                round(v["Dmax"],  4),       round(v["Sigma_Dmax"],  4),
                round(v["D95"],   4),       round(v["Sigma_D95"],   4),
                round(v["D5"],    4),       round(v["Sigma_D5"],    4),
            ]
            for ci, val in enumerate(vals, start=1):
                c = ws_res.cell(row=ridx, column=ci, value=val)
                c.fill = fill_row
                c.border = border
                c.alignment = left_al if ci == 1 else right_al

        # ═══════════════════════════════════════════════════════════════════
        # HOJA 3: Componentes de dosis
        # ═══════════════════════════════════════════════════════════════════
        comp_data = self.report.get("CompDoses", {})
        if comp_data:
            ws_comp = wb.create_sheet("Componentes de dosis")
            COMPS_XL = [
                ("Boro",  "Boro"),
                ("Fstn",  "Neut. Rápidos"),
                ("Thn",   "Neut. Térmicos"),
                ("Gamma", "Gamma"),
            ]
            SUB_XL = ["Promedio", "Mínima", "Máxima"]

            # Fila 1: título
            total_cols = 1 + len(COMPS_XL) * 3 + 3
            _hdr_cell(ws_comp, 1, 1, "Componentes de Dosis por Órgano (Dosis Física, sin RBE/CBE)")
            ws_comp.merge_cells(f"A1:{get_column_letter(total_cols)}1")

            # Fila 2: títulos de grupo
            ws_comp.column_dimensions["A"].width = 18
            ws_comp.cell(row=2, column=1, value="Órgano").font = hdr_font
            ws_comp.cell(row=2, column=1).fill = hdr_fill
            ws_comp.cell(row=2, column=1).border = border
            ws_comp.cell(row=2, column=1).alignment = center
            ws_comp.merge_cells(f"A2:A3")

            for gi, (_, label) in enumerate(COMPS_XL):
                sc = 2 + gi * 3
                _hdr_cell(ws_comp, 2, sc, label)
                ws_comp.merge_cells(
                    f"{get_column_letter(sc)}2:{get_column_letter(sc+2)}2"
                )

            # Total
            tc = 2 + len(COMPS_XL) * 3
            _hdr_cell(ws_comp, 2, tc, "Total")
            ws_comp.merge_cells(
                f"{get_column_letter(tc)}2:{get_column_letter(tc+2)}2"
            )

            # Fila 3: sub-columnas
            for gi in range(len(COMPS_XL) + 1):
                for si, s_lbl in enumerate(SUB_XL):
                    cc = 2 + gi * 3 + si
                    _hdr_cell(ws_comp, 3, cc, s_lbl, fill=hdr2_fill, font=hdr2_font)

            for ridx, organ in enumerate(sorted(comp_data.keys()), start=4):
                cd = comp_data[organ]
                fill_row = PatternFill("solid", fgColor="F5F7FA" if ridx % 2 == 0 else "FFFFFF")
                c = ws_comp.cell(row=ridx, column=1, value=organ)
                c.fill = fill_row; c.border = border; c.alignment = left_al

                col_idx = 2
                arrays = {}
                for comp_key, _ in COMPS_XL:
                    arr = np.array(cd.get(comp_key, []), dtype=float)
                    arrays[comp_key] = arr[np.isfinite(arr)]

                valid = [arrays[k] for k in arrays if arrays[k].size > 0]
                ref_size = valid[0].size if valid else 1
                total_arr = np.zeros(ref_size)
                for arr in valid:
                    n = min(arr.size, ref_size)
                    total_arr[:n] += arr[:n]

                for comp_key, _ in COMPS_XL:
                    arr = arrays[comp_key]
                    if arr.size > 0:
                        vals_c = (float(np.nanmean(arr)), float(np.nanmin(arr)), float(np.nanmax(arr)))
                    else:
                        vals_c = (0.0, 0.0, 0.0)
                    for val in vals_c:
                        cx = ws_comp.cell(row=ridx, column=col_idx, value=round(val, 4))
                        cx.fill = fill_row; cx.border = border; cx.alignment = right_al
                        col_idx += 1

                # Total
                t_vals = (float(np.nanmean(total_arr)), float(np.nanmin(total_arr)), float(np.nanmax(total_arr)))
                for val in t_vals:
                    cx = ws_comp.cell(row=ridx, column=col_idx, value=round(val, 4))
                    cx.fill = fill_row; cx.border = border; cx.alignment = right_al
                    col_idx += 1

        wb.save(fname)
        QtWidgets.QMessageBox.information(self, "Excel", f"Guardado en:\n{fname}")

    def _export_pdf(self):
        try:
            from reportlab.lib.pagesizes import A4
            from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, Image
            from reportlab.lib.styles import getSampleStyleSheet
            from reportlab.lib import colors
            from reportlab.lib.utils import ImageReader
        except Exception:
            QtWidgets.QMessageBox.warning(self, "PDF no disponible", "Instalá reportlab para exportar PDF.")
            return

        sugerido = Path.home()/("resultados_ratmaster_"+datetime.now().strftime("%Y%m%dT%H%M%SZ")+".pdf")
        fname, _ = QtWidgets.QFileDialog.getSaveFileName(self, "Guardar PDF", str(sugerido), "PDF (*.pdf)")
        if not fname:
            return
        out = Path(fname)

        styles = getSampleStyleSheet()
        story = []

        
        # Helpers de layout
        PAGE_W, PAGE_H = A4
        USABLE_W = PAGE_W - 36 - 36  # márgenes
        def fit_colwidths(widths, maxw=USABLE_W):
            total = sum(widths)
            if total <= maxw:
                return widths
            scale = maxw / total
            return [w*scale for w in widths]
# ===== Cabecera / metadatos =====
        meta = self.report.get("meta", {})
        story.append(Paragraph("Reporte dosimétrico de la irradiación", styles["Title"]))
        story.append(Spacer(1, 6))
        _MESES_ES = ["enero", "febrero", "marzo", "abril", "mayo", "junio", 
             "julio", "agosto", "septiembre", "octubre", "noviembre", "diciembre"]
        try:
            # Obtenemos el string de la fecha
            date_str = meta.get("date", "")
            if date_str:
                _dt = datetime.fromisoformat(date_str).astimezone()
                # Construcción limpia con f-string
                _fecha_fmt = f"{_dt.day} de {_MESES_ES[_dt.month-1]} de {_dt.year}, {_dt.strftime('%H:%M:%S')}"
            else:
                _fecha_fmt = "Fecha no disponible"
        except Exception:
            # Si falla el parseo, devolvemos el valor original o un fallback
            _fecha_fmt = meta.get("date", "Error en formato de fecha")
        md = [
            ["Fecha (UTC)", _fecha_fmt],
            ["Modo", "Restricciones" if meta.get("mode")=="constraints" else "Tiempo fijo"],
            ["Tipo de dosis", self.dose_mode_text],
            ["Protocolo de boro", _resolve_boro_protocol_name(meta.get("boro_protocol_name", "Manual"))],
            ["Origen del boro", format_boro_origin(meta)],
            ["Tiempo", format_time(meta.get("time",0), meta.get("time_err",0), "s")],
            ["Flujo neutrónico", format_scientific_value_uncertainty(meta.get("spnd",0), meta.get("spnd_abs_err",0), "n/cm²·s")],
        ]
        if meta.get("mode")=="constraints" and meta.get("chosen_constraint"):
            ch = meta["chosen_constraint"]
            org = ch.get("org","?"); typ = ch.get("type","?")
            _pdf_unit = dose_axis_unit(self.use_bio, self.is_isoe)
            lim = ch.get("limit_value", ch.get("limit","?"))
            ach = ch.get("achieved_value", None)
            tval = ch.get("time_computed", meta.get("time", None))
            try:
                ach_txt = f"{float(ach):.3f}" if ach is not None else "?"
            except Exception:
                ach_txt = str(ach)
            extra_t = f", t={tval:.3f}s" if isinstance(tval,(int,float)) else ""
            md.append(["Restricción limitante", f"{org} - {typ} (límite {lim} {_pdf_unit}, logrado {ach_txt} {_pdf_unit}{extra_t})"])

        constraints_used = meta.get("constraints_used") or []
        if meta.get("mode") == "constraints" and constraints_used:
            md.append(["Cantidad de restricciones", str(len(constraints_used))])
        t_meta = Table([["Campo","Valor"]]+md, colWidths=(150, 370))
        t_meta.setStyle(TableStyle([
            ("BACKGROUND",(0,0),(-1,0), colors.HexColor("#f0f0f0")),
            ("GRID",(0,0),(-1,-1), 0.25, colors.grey),
            ("FONTNAME",(0,0),(-1,0), "Helvetica-Bold"),
            ("VALIGN",(0,0),(-1,-1), "MIDDLE"),
        ]))
        story.append(t_meta)
        story.append(Spacer(1, 12))

        # ===== Flujo desde corrientes SPND (si aplica) =====
        snap = meta.get("spnd_from_currents")
        if snap:
            story.append(Paragraph("Medición de corrientes SPND", styles["Heading2"]))
            rows_snap = snap.get("rows", [])
            data_spnd = [["Detector", "Usar", "I medida (pA)", "± I (pA)"]]
            for rd in rows_snap:
                data_spnd.append([
                    rd.get("detector", "—"),
                    "Sí" if rd.get("usar") else "No",
                    rd.get("I_meas", "—"),
                    rd.get("I_sigma", "—"),
                ])
            t_spnd = Table(data_spnd, colWidths=fit_colwidths([130, 50, 130, 130]))
            t_spnd.setStyle(TableStyle([
                ("BACKGROUND",(0,0),(-1,0), colors.HexColor("#f0f0f0")),
                ("GRID",(0,0),(-1,-1), 0.25, colors.grey),
                ("FONTNAME",(0,0),(-1,0), "Helvetica-Bold"),
                ("VALIGN",(0,0),(-1,-1), "MIDDLE"),
            ]))
            story.append(t_spnd)
            story.append(Spacer(1, 6))
            cip_md = [
                ["Referencia", snap.get("ref_detector", "—")],
                ["Sensibilidad ref.", snap.get("sens_txt", "—")],
                ["CIP monitoreo", snap.get("cip_mon", "—")],
                ["CIP irradiación", snap.get("cip_irr", "—")],
                ["CIP pert", snap.get("cip_pert", "—")],
                ["I media equiv. (pA)", snap.get("Imean_txt", "—")],
                ["I irradiación (pA)", snap.get("Iirr_txt", "—")],
                ["Flujo resultante", snap.get("phi_txt", "—")],
                ["Incertidumbre relativa", snap.get("break_txt", "—")],
            ]
            t_cip = Table([["Parámetro", "Valor"]] + cip_md, colWidths=(200, 320))
            t_cip.setStyle(TableStyle([
                ("BACKGROUND",(0,0),(-1,0), colors.HexColor("#f0f0f0")),
                ("GRID",(0,0),(-1,-1), 0.25, colors.grey),
                ("FONTNAME",(0,0),(-1,0), "Helvetica-Bold"),
                ("VALIGN",(0,0),(-1,-1), "MIDDLE"),
            ]))
            story.append(t_cip)
            story.append(Spacer(1, 12))

        constraints_used = meta.get("constraints_used") or []
        if meta.get("mode") == "constraints" and constraints_used:
            story.append(Paragraph("Restricciones consideradas", styles["Heading2"]))
            _cons_unit = dose_axis_unit(self.use_bio, self.is_isoe)
            data_cons = [["Órgano", "Restricción", f"Límite ({_cons_unit})"]]
            for c in constraints_used:
                org = str(c.get("org", "?"))
                disp = str(c.get("display") or c.get("metric") or "?")
                lim = c.get("limit_value", "")
                try:
                    lim_txt = f"{float(lim):.3f} {_cons_unit}"
                except Exception:
                    lim_txt = str(lim)
                data_cons.append([org, disp, lim_txt])
            t_cons = Table(data_cons, colWidths=fit_colwidths([130, 250, 140]))
            t_cons.setStyle(TableStyle([
                ("BACKGROUND",(0,0),(-1,0), colors.HexColor("#f0f0f0")),
                ("GRID",(0,0),(-1,-1), 0.25, colors.grey),
                ("FONTNAME",(0,0),(-1,0), "Helvetica-Bold"),
                ("VALIGN",(0,0),(-1,-1), "MIDDLE"),
            ]))
            story.append(t_cons)
            story.append(Spacer(1, 12))

        # ===== Parámetros (tabla) =====
        B   = self.report.get("ParamsUsed", {}).get("B", [])
        Berr= self.report.get("ParamsUsed", {}).get("B_err", [])
        CBE = self.report.get("ParamsUsed", {}).get("CBE", [])
        RBE = self.report.get("ParamsUsed", {}).get("RBE", [])
        
        show_bio_factors = bool(self.use_bio and not getattr(self, "is_isoe", False))

        # Cabecera dinámica
        if show_bio_factors:
            data_params = [["Órgano", "B [ppm] ±", "CBE", "RBE"]]
        else:
            data_params = [["Órgano", "B [ppm] ±"]]
        for j, name in enumerate(ORG_ORDER):
            b  = B[j] if j < len(B) else 0.0
            be = Berr[j] if j < len(Berr) else 0.0
            cbe= CBE[j] if j < len(CBE) else 1.0
            rbe= RBE[j] if j < len(RBE) else 1.0
            row = [name, f"{b:.6g} ± {be:.6g}"]
            if show_bio_factors:
                row += [f"{cbe:.6g}", f"{rbe:.6g}"]
            data_params.append(row)
            
        cols_params = [120, 160] + ([120, 120] if show_bio_factors else [])
        t_params = Table(data_params, colWidths=fit_colwidths(cols_params))
        t_params.setStyle(TableStyle([
            ("BACKGROUND",(0,0),(-1,0), colors.HexColor("#f0f0f0")),
            ("GRID",(0,0),(-1,-1), 0.25, colors.grey),
            ("FONTNAME",(0,0),(-1,0), "Helvetica-Bold"),
            ("ALIGN",(1,1),(-1,-1), "RIGHT"),
        ]))
        story.append(Paragraph("Parámetros de entrada", styles["Heading2"]))
        story.append(t_params)
        story.append(Spacer(1, 12))

        if getattr(self, "is_isoe", False):
            meta_isoe = self.report.get("meta", {})
            params = meta_isoe.get("isoe_params_dict") or {}
            t0_map = meta_isoe.get("isoe_t0_map") or {}
            preset_name = meta_isoe.get("isoe_preset_name", "Manual")
            story.append(Paragraph("Parámetros usados para Dosis Isoefectiva (Photon IsoE): Gy(IsoE)", styles["Heading2"]))
            data_iso_meta = [["Campo", "Valor"], ["Preset IsoE", str(preset_name)]]
            t_iso_meta = Table(data_iso_meta, colWidths=fit_colwidths([160, 260]))
            t_iso_meta.setStyle(TableStyle([
                ("BACKGROUND",(0,0),(-1,0), colors.HexColor("#f0f0f0")),
                ("GRID",(0,0),(-1,-1), 0.25, colors.grey),
                ("FONTNAME",(0,0),(-1,0), "Helvetica-Bold"),
            ]))
            story.append(t_iso_meta)
            story.append(Spacer(1, 8))
            if params:
                ordered = ["aR","bR","GR","aB","bB","aTh","bTh","aFn","bFn","aG","bG"]
                data_iso = [["Parámetro", "Valor"]] + [[k, f"{float(params[k]):.6g}"] for k in ordered if k in params]
                t_iso = Table(data_iso, colWidths=fit_colwidths([160, 180]))
                t_iso.setStyle(TableStyle([
                    ("BACKGROUND",(0,0),(-1,0), colors.HexColor("#f0f0f0")),
                    ("GRID",(0,0),(-1,-1), 0.25, colors.grey),
                    ("FONTNAME",(0,0),(-1,0), "Helvetica-Bold"),
                    ("ALIGN",(1,1),(-1,-1), "RIGHT"),
                ]))
                story.append(t_iso)
                story.append(Spacer(1, 8))
            if t0_map:
                ordered_t0 = ["Boro","Fstn","Thn","Gamma","R"]
                data_t0 = [["t0", "Valor (s)"]] + [[k, f"{float(t0_map[k]):.6g}"] for k in ordered_t0 if k in t0_map]
                t_t0 = Table(data_t0, colWidths=fit_colwidths([160, 180]))
                t_t0.setStyle(TableStyle([
                    ("BACKGROUND",(0,0),(-1,0), colors.HexColor("#f0f0f0")),
                    ("GRID",(0,0),(-1,-1), 0.25, colors.grey),
                    ("FONTNAME",(0,0),(-1,0), "Helvetica-Bold"),
                    ("ALIGN",(1,1),(-1,-1), "RIGHT"),
                ]))
                story.append(t_t0)
                story.append(Spacer(1, 12))

        # ===== Métricas (tabla) =====
        head = ("Métricas de Dosis Isoefectiva (Photon IsoE): Gy(IsoE)" if getattr(self,"is_isoe",False) else ("Métricas de Dosis Equivalente pesada por RBE: Gy(RBE)" if self.use_bio else "Métricas de Dosis Física: Gy"))
        data_metrics = [["Órgano", "Dosis Promedio ± σ", "Dosis Mínima ± σ", "Dosis Máxima ± σ", "D95 ± σ", "D5 ± σ"]]
        for key in sorted(self.dsum.keys()):
            v = self.dsum[key]
            # [FIX] Antes: f"{v['Dmean']:.2f} ± {v['Sigma_Dmean']:.2f}" con
            # decimales fijos sin relación con la cifra significativa real
            # de cada sigma. Ahora usa la misma regla de redondeo que el
            # resto de la app (ver ui/formatters.format_value_uncertainty).
            row = [
                key,
                format_value_uncertainty(v['Dmean'], v['Sigma_Dmean']),
                format_value_uncertainty(v['Dmin'],  v['Sigma_Dmin']),
                format_value_uncertainty(v['Dmax'],  v['Sigma_Dmax']),
                format_value_uncertainty(v['D95'],   v['Sigma_D95']),
                format_value_uncertainty(v['D5'],    v['Sigma_D5']),
            ]
            data_metrics.append(row)
        cols_metrics = [90, 120, 120, 120, 120, 120]
        t_metrics = Table(data_metrics, colWidths=fit_colwidths(cols_metrics))
        t_metrics.setStyle(TableStyle([
            ("BACKGROUND",(0,0),(-1,0), colors.HexColor("#f0f0f0")),
            ("GRID",(0,0),(-1,-1), 0.25, colors.grey),
            ("FONTNAME",(0,0),(-1,0), "Helvetica-Bold"),
            ("ALIGN",(1,1),(-1,-1), "RIGHT"),
        ]))
        story.append(Paragraph(head, styles["Heading2"]))
        story.append(t_metrics)
        story.append(Spacer(1, 12))

        # ===== DVH como imagen (extendido a 0 Gy) =====
        buf = io.BytesIO()
        fig = Figure(figsize=(6,3), dpi=150)
        FigureCanvasAgg(fig)   # necesario para que fig.savefig() funcione sin pantalla
        ax = fig.add_subplot(111)
        ax.set_xlabel(f"Dosis ({dose_axis_unit(self.use_bio, self.is_isoe)})")
        ax.set_ylabel("Volumen (%)")
        ax.set_title("DVH — Dosis Isoefectiva (Photon IsoE)" if self.is_isoe else ("DVH — Dosis Equivalente pesada por RBE" if self.use_bio else "DVH — Dosis Física"))
        voxmap = self.report.get("IsoVoxel" if self.is_isoe else ("BioVoxel" if self.use_bio else "PhysVoxel"), {})
        _keys_pdf = list(voxmap.keys())
        _cmap_pdf = {k: ORGAN_COLORS[i % len(ORGAN_COLORS)] for i, k in enumerate(_keys_pdf)}
        for key, vox in voxmap.items():
            arr = np.array(vox, dtype=float)
            if arr.size == 0:
                continue
            s, vol = build_dvh(arr)
            s, vol = dvh_extend_to_zero(s, vol)
            if s.size > 0:
                ax.plot(s, vol, label=key, color=_cmap_pdf[key])
        ax.set_xlim(left=0); ax.set_ylim(0, 105)
        ax.grid(True, which='major', color='#E0E0E0', linewidth=0.8, alpha=0.9)
        ax.grid(True, which='minor', color='#EEEEEE', linewidth=0.4, alpha=0.7)
        ax.minorticks_on()
        ax.set_axisbelow(True)
        ax.legend(fontsize=7)
        fig.tight_layout()
        fig.savefig(buf, format="png"); buf.seek(0)

        story.append(Paragraph("DVH", styles["Heading2"]))
        buf.seek(0)  # asegurar que el puntero está al inicio
        img = RLImage(buf)
        try:
            # Limitar al ancho útil manteniendo proporción
            img._restrictSize(USABLE_W, 240)
        except Exception:
            pass
        story.append(img)

        # ===== Tabla de Componentes de Dosis =====
        comp_data = self.report.get("CompDoses", {})
        if comp_data:
            story.append(Spacer(1, 14))
            story.append(Paragraph("Componentes de Dosis por Órgano (Dosis Física)", styles["Heading2"]))
            story.append(Paragraph(
                "Dmean y Dmax de cada componente, ya multiplicados por el tiempo de irradiación. "
                "Los factores RBE/CBE no están aplicados. "
                "La columna % indica la fracción respecto al Dmean total del órgano.",
                styles["Normal"]
            ))
            story.append(Spacer(1, 6))

            COMPS = [("Boro", "D_Boro"), ("Fstn", "D_Ráp."), ("Thn", "D_Tér."), ("Gamma", "D_Gamma")]
            organs_sorted = sorted(comp_data.keys())

            # Cabecera en dos líneas fusionadas con SPAN
            header_row1 = ["Órgano"] + [f"{lbl}" for _, lbl in COMPS] + ["Total"]
            header_row2 = [""] + ["Dmean / Dmax / %" for _ in COMPS] + ["Dmean / Dmax"]

            data_comp = [header_row1, header_row2]
            for organ in organs_sorted:
                cd = comp_data[organ]
                arrays = {}
                for comp_key, _ in COMPS:
                    arr = np.array(cd.get(comp_key, []), dtype=float)
                    arrays[comp_key] = arr[np.isfinite(arr)]

                total_arr = np.zeros(max((arrays[k].size for k in arrays), default=1))
                for k in arrays:
                    if arrays[k].size == total_arr.size:
                        total_arr = total_arr + arrays[k]
                    elif arrays[k].size > 0:
                        # Tamaños distintos: sumar lo que se pueda
                        n = min(arrays[k].size, total_arr.size)
                        total_arr[:n] += arrays[k][:n]
                total_mean = float(np.nanmean(total_arr)) if total_arr.size > 0 else 0.0
                total_max  = float(np.nanmax(total_arr))  if total_arr.size > 0 else 0.0

                row = [organ]
                for comp_key, _ in COMPS:
                    arr = arrays[comp_key]
                    if arr.size > 0:
                        dmean = float(np.nanmean(arr))
                        dmax  = float(np.nanmax(arr))
                    else:
                        dmean = 0.0; dmax = 0.0
                    pct = (dmean / total_mean * 100.0) if total_mean > 0 else 0.0
                    row.append(f"{dmean:.3f} / {dmax:.3f}\n{pct:.1f}%")
                row.append(f"{total_mean:.3f} / {total_max:.3f}")
                data_comp.append(row)

            n_cols = 1 + len(COMPS) + 1
            col_w_org = 75
            col_w_comp = (USABLE_W - col_w_org - 65) / len(COMPS)
            col_widths_comp = [col_w_org] + [col_w_comp] * len(COMPS) + [65]

            t_comp = Table(data_comp, colWidths=fit_colwidths(col_widths_comp), repeatRows=2)
            style_comp = TableStyle([
                # Cabecera fila 0
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#455A64")),
                ("TEXTCOLOR",  (0, 0), (-1, 0), colors.white),
                ("FONTNAME",   (0, 0), (-1, 0), "Helvetica-Bold"),
                ("ALIGN",      (0, 0), (-1, 0), "CENTER"),
                # Cabecera fila 1
                ("BACKGROUND", (0, 1), (-1, 1), colors.HexColor("#f0f0f0")),
                ("FONTNAME",   (0, 1), (-1, 1), "Helvetica-Bold"),
                ("ALIGN",      (0, 1), (-1, 1), "CENTER"),
                # Datos
                ("ALIGN",      (1, 2), (-1, -1), "CENTER"),
                ("ALIGN",      (0, 2), (0, -1),  "LEFT"),
                ("VALIGN",     (0, 0), (-1, -1), "MIDDLE"),
                ("GRID",       (0, 0), (-1, -1), 0.25, colors.grey),
                ("ROWBACKGROUNDS", (0, 2), (-1, -1), [colors.white, colors.HexColor("#F5F7FA")]),
                ("FONTSIZE",   (0, 0), (-1, -1), 7.5),
            ])
            t_comp.setStyle(style_comp)
            story.append(t_comp)
            story.append(Spacer(1, 12))

        # ==== Guardar ====
        doc = SimpleDocTemplate(str(out), pagesize=A4, leftMargin=36, rightMargin=36, topMargin=36, bottomMargin=36)
        doc.build(story)
        QtWidgets.QMessageBox.information(self, "PDF", f"Guardado en:\n{out}")
